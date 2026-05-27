# -*- coding: utf-8 -*-
"""
2024 高教社杯数学建模 C 题：问题一 MILP 完整求解脚本

功能：
1. 读取 附件1.xlsx、附件2.xlsx；
2. 建立 2024-2030 年农作物种植混合整数线性规划模型；
3. 同时求解问题一的两个情形：
   - result1_1：超过预期销售量部分滞销浪费；
   - result1_2：超过预期销售量部分按 50% 售价销售；
4. 写回 附件3/result1_1.xlsx、附件3/result1_2.xlsx 模板。

默认项目结构：
C题/
├─ 附件1.xlsx
├─ 附件2.xlsx
├─ 附件3/
│  ├─ result1_1.xlsx
│  └─ result1_2.xlsx
└─ solve_problem1_milp_final.py

运行：
python solve_problem1_milp_final.py

指定路径运行：
python solve_problem1_milp_final.py --base "D:\\你的路径\\C题"

说明：
- 这是 MILP，不是 baseline/随机/贪心。
- 使用 x 连续变量表示面积，z 二进制变量表示是否种植。
- 允许同一地块同一季合种多个作物，同时用“最小种植面积”和“每季最多作物数”控制碎片化。
- 如果求解时间太长，可以适当调大 --gap 或调小 --time-limit。
"""

from __future__ import annotations

import argparse
import math
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
import pulp
from openpyxl import Workbook, load_workbook

YEARS = list(range(2024, 2031))
SEASONS = [1, 2]


def clean_text(x: Any) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()


def is_number(x: Any) -> bool:
    try:
        if pd.isna(x):
            return False
        float(x)
        return True
    except Exception:
        return False


def safe_value(x: Any, default: float = 0.0) -> float:
    if x is None:
        return default
    try:
        v = float(x)
        if math.isnan(v):
            return default
        return v
    except Exception:
        return default


def parse_price(value: Any) -> float:
    s = clean_text(value)
    if not s:
        raise ValueError("销售单价为空，无法解析")
    if "-" in s:
        a, b = s.split("-", 1)
        return (float(a) + float(b)) / 2.0
    return float(s)


def stat_season_to_model_season(s: str) -> int:
    s = clean_text(s)
    return 1 if s in ["单季", "第一季"] else 2


def load_data(base_dir: Path, att1_name: str, att2_name: str) -> Dict[str, Any]:
    att1 = base_dir / att1_name
    att2 = base_dir / att2_name
    if not att1.exists():
        raise FileNotFoundError(f"找不到 {att1}")
    if not att2.exists():
        raise FileNotFoundError(f"找不到 {att2}")

    land_df = pd.read_excel(att1, sheet_name="乡村的现有耕地")
    crop_df = pd.read_excel(att1, sheet_name="乡村种植的农作物")
    plant_df = pd.read_excel(att2, sheet_name="2023年的农作物种植情况")
    stat_df = pd.read_excel(att2, sheet_name="2023年统计的相关数据")

    lands: List[Dict[str, Any]] = []
    for _, row in land_df.iterrows():
        plot = clean_text(row.get("地块名称"))
        land_type = clean_text(row.get("地块类型"))
        area = row.get("地块面积/亩")
        if not plot or not land_type or not is_number(area):
            continue
        lands.append({"plot": plot, "land_type": land_type, "area": float(area)})
    land_by_plot = {r["plot"]: r for r in lands}

    crop_df = crop_df[pd.to_numeric(crop_df["作物编号"], errors="coerce").notna()].copy()
    crop_df["作物编号"] = crop_df["作物编号"].astype(int)
    crops: List[Dict[str, Any]] = []
    for _, row in crop_df.iterrows():
        cid = int(row["作物编号"])
        crops.append({
            "crop_id": cid,
            "crop_name": clean_text(row["作物名称"]),
            "crop_type": clean_text(row["作物类型"]),
        })
    crop_by_id = {c["crop_id"]: c for c in crops}
    crop_name_by_id = {c["crop_id"]: c["crop_name"] for c in crops}
    crop_id_by_name = {c["crop_name"]: c["crop_id"] for c in crops}
    bean_ids = {c["crop_id"] for c in crops if "豆类" in c["crop_type"]}

    stat_df = stat_df[pd.to_numeric(stat_df["作物编号"], errors="coerce").notna()].copy()
    stat_df["作物编号"] = stat_df["作物编号"].astype(int)
    stats: Dict[Tuple[int, str, str], Dict[str, float]] = {}
    for _, row in stat_df.iterrows():
        cid = int(row["作物编号"])
        land_type = clean_text(row["地块类型"])
        stat_season = clean_text(row["种植季次"])
        stats[(cid, land_type, stat_season)] = {
            "yield": float(row["亩产量/斤"]),
            "cost": float(row["种植成本/(元/亩)"]),
            "price": parse_price(row["销售单价/(元/斤)"]),
        }

    # 题目附件说明：智慧大棚第一季蔬菜的亩产量、成本、价格与普通大棚相同，统计表省略。
    for cid in range(17, 35):
        src = (cid, "普通大棚", "第一季")
        dst = (cid, "智慧大棚", "第一季")
        if src in stats and dst not in stats:
            stats[dst] = dict(stats[src])

    plant_df["种植地块"] = plant_df["种植地块"].ffill()
    plant_df = plant_df[pd.to_numeric(plant_df["作物编号"], errors="coerce").notna()].copy()
    plant_df["作物编号"] = plant_df["作物编号"].astype(int)

    demand = {cid: 0.0 for cid in crop_by_id}
    plant_2023_by_plot_season: Dict[Tuple[str, int], set[int]] = {}
    bean_2023_flag = {land["plot"]: 0 for land in lands}

    for _, row in plant_df.iterrows():
        plot = clean_text(row["种植地块"])
        if plot not in land_by_plot:
            continue
        cid = int(row["作物编号"])
        area = float(row["种植面积/亩"])
        stat_season = clean_text(row["种植季次"])
        land_type = land_by_plot[plot]["land_type"]
        key = (cid, land_type, stat_season)
        if key not in stats:
            raise KeyError(f"统计参数缺失：作物 {cid}，地块类型 {land_type}，季次 {stat_season}")
        demand[cid] += area * stats[key]["yield"]
        season = stat_season_to_model_season(stat_season)
        plant_2023_by_plot_season.setdefault((plot, season), set()).add(cid)
        if cid in bean_ids:
            bean_2023_flag[plot] = 1

    return {
        "lands": lands,
        "land_by_plot": land_by_plot,
        "crops": crops,
        "crop_by_id": crop_by_id,
        "crop_name_by_id": crop_name_by_id,
        "crop_id_by_name": crop_id_by_name,
        "bean_ids": bean_ids,
        "stats": stats,
        "demand": demand,
        "plant_2023_by_plot_season": plant_2023_by_plot_season,
        "bean_2023_flag": bean_2023_flag,
    }


def feasible_crops_for_plot(land_type: str, season: int) -> List[Tuple[int, str]]:
    """返回某地块类型、某季允许种植的作物编号及其对应统计季次。"""
    if land_type in ["平旱地", "梯田", "山坡地"]:
        if season == 1:
            return [(cid, "单季") for cid in range(1, 16)]
        return []

    if land_type == "水浇地":
        if season == 1:
            # 单季水稻，或两季蔬菜的第一季。大白菜/白萝卜/红萝卜只能第二季。
            return [(16, "单季")] + [(cid, "第一季") for cid in range(17, 35)]
        if season == 2:
            return [(cid, "第二季") for cid in [35, 36, 37]]
        return []

    if land_type == "普通大棚":
        if season == 1:
            return [(cid, "第一季") for cid in range(17, 35)]
        if season == 2:
            return [(cid, "第二季") for cid in range(38, 42)]
        return []

    if land_type == "智慧大棚":
        if season == 1:
            return [(cid, "第一季") for cid in range(17, 35)]
        if season == 2:
            return [(cid, "第二季") for cid in range(17, 35)]
        return []

    return []


def max_crops_allowed(land_type: str, season: int) -> int:
    """控制同一地块同一季的作物数量，避免方案过碎。"""
    if land_type in ["平旱地", "梯田", "山坡地"]:
        return 2
    if land_type == "水浇地":
        return 3 if season == 1 else 1
    if land_type == "普通大棚":
        return 3 if season == 1 else 1
    if land_type == "智慧大棚":
        return 3
    return 1


def min_area_for_land(land_type: str, area: float, min_open: float, min_greenhouse: float) -> float:
    if "大棚" in land_type:
        return min(min_greenhouse, area)
    return min(min_open, area)


def build_records(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    rid = 0
    stats = data["stats"]
    for year in YEARS:
        for land in data["lands"]:
            plot = land["plot"]
            land_type = land["land_type"]
            area = land["area"]
            for season in SEASONS:
                for cid, stat_season in feasible_crops_for_plot(land_type, season):
                    key = (cid, land_type, stat_season)
                    if key not in stats:
                        continue
                    st = stats[key]
                    records.append({
                        "rid": rid,
                        "year": year,
                        "plot": plot,
                        "land_type": land_type,
                        "plot_area": area,
                        "season": season,
                        "crop_id": cid,
                        "stat_season": stat_season,
                        "yield": st["yield"],
                        "cost": st["cost"],
                        "price": st["price"],
                    })
                    rid += 1
    return records


def solve_one_case(
    data: Dict[str, Any],
    discount_rate: float,
    case_name: str,
    time_limit: int,
    gap: float,
    min_open: float,
    min_greenhouse: float,
    solver_msg: bool,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    records = build_records(data)
    rid_to_rec = {r["rid"]: r for r in records}

    by_year_plot_season: Dict[Tuple[int, str, int], List[int]] = {}
    by_year_plot_season_crop: Dict[Tuple[int, str, int, int], int] = {}
    by_year_crop: Dict[Tuple[int, int], List[int]] = {}
    by_plot_year_season_crop: Dict[Tuple[str, int, int, int], int] = {}

    for r in records:
        rid = r["rid"]
        key_slot = (r["year"], r["plot"], r["season"])
        by_year_plot_season.setdefault(key_slot, []).append(rid)
        by_year_plot_season_crop[(r["year"], r["plot"], r["season"], r["crop_id"])] = rid
        by_year_crop.setdefault((r["year"], r["crop_id"]), []).append(rid)
        by_plot_year_season_crop[(r["plot"], r["year"], r["season"], r["crop_id"])] = rid

    model = pulp.LpProblem(f"C_problem1_{case_name}", pulp.LpMaximize)

    x = {rid: pulp.LpVariable(f"x_{rid}", lowBound=0, cat="Continuous") for rid in rid_to_rec}
    z = {rid: pulp.LpVariable(f"z_{rid}", lowBound=0, upBound=1, cat="Binary") for rid in rid_to_rec}
    sale = {rid: pulp.LpVariable(f"sale_{rid}", lowBound=0, cat="Continuous") for rid in rid_to_rec}

    # 面积-二进制变量关联；被选作物至少达到最小面积。
    for rid, rec in rid_to_rec.items():
        area = rec["plot_area"]
        lb = min_area_for_land(rec["land_type"], area, min_open, min_greenhouse)
        model += x[rid] <= area * z[rid], f"link_upper_{rid}"
        model += x[rid] >= lb * z[rid], f"link_lower_{rid}"

        production = rec["yield"] * x[rid]
        model += sale[rid] <= production, f"sale_le_production_{rid}"

    # 年度预期销售量限制：每年每种作物正常价格销售量 <= 2023 年预期销售量。
    for year in YEARS:
        for cid, demand in data["demand"].items():
            rids = by_year_crop.get((year, cid), [])
            if rids:
                model += pulp.lpSum(sale[rid] for rid in rids) <= demand, f"demand_{year}_{cid}"

    # 土地/季节结构约束。
    water_two_mode: Dict[Tuple[int, str], pulp.LpVariable] = {}

    for year in YEARS:
        for land in data["lands"]:
            plot = land["plot"]
            land_type = land["land_type"]
            area = land["area"]

            if land_type in ["平旱地", "梯田", "山坡地"]:
                rids = by_year_plot_season.get((year, plot, 1), [])
                model += pulp.lpSum(x[rid] for rid in rids) == area, f"dry_area_{year}_{plot}"
                model += pulp.lpSum(z[rid] for rid in rids) <= max_crops_allowed(land_type, 1), f"dry_max_crop_{year}_{plot}"

            elif land_type == "水浇地":
                mode = pulp.LpVariable(f"water_two_{year}_{plot}", lowBound=0, upBound=1, cat="Binary")
                water_two_mode[(year, plot)] = mode
                rice_rid = by_year_plot_season_crop.get((year, plot, 1, 16))
                if rice_rid is None:
                    raise RuntimeError(f"水浇地 {plot} 缺少水稻可选记录")

                first_veg_rids = [
                    rid for rid in by_year_plot_season.get((year, plot, 1), [])
                    if 17 <= rid_to_rec[rid]["crop_id"] <= 34
                ]
                second_veg_rids = by_year_plot_season.get((year, plot, 2), [])

                # mode=0：种一季水稻，第一季水稻面积=area，第二季不种。
                # mode=1：种两季蔬菜，第一季蔬菜面积=area，第二季蔬菜面积=area。
                model += x[rice_rid] == area * (1 - mode), f"water_rice_area_{year}_{plot}"
                model += z[rice_rid] == 1 - mode, f"water_rice_z_{year}_{plot}"
                model += pulp.lpSum(x[rid] for rid in first_veg_rids) == area * mode, f"water_first_veg_area_{year}_{plot}"
                model += pulp.lpSum(x[rid] for rid in second_veg_rids) == area * mode, f"water_second_veg_area_{year}_{plot}"
                model += pulp.lpSum(z[rid] for rid in first_veg_rids) <= max_crops_allowed(land_type, 1) * mode, f"water_first_max_{year}_{plot}"
                model += pulp.lpSum(z[rid] for rid in second_veg_rids) <= max_crops_allowed(land_type, 2) * mode, f"water_second_max_{year}_{plot}"

            elif land_type == "普通大棚":
                for season in SEASONS:
                    rids = by_year_plot_season.get((year, plot, season), [])
                    model += pulp.lpSum(x[rid] for rid in rids) == area, f"normal_greenhouse_area_{year}_{plot}_{season}"
                    model += pulp.lpSum(z[rid] for rid in rids) <= max_crops_allowed(land_type, season), f"normal_greenhouse_max_{year}_{plot}_{season}"

            elif land_type == "智慧大棚":
                for season in SEASONS:
                    rids = by_year_plot_season.get((year, plot, season), [])
                    model += pulp.lpSum(x[rid] for rid in rids) == area, f"smart_greenhouse_area_{year}_{plot}_{season}"
                    model += pulp.lpSum(z[rid] for rid in rids) <= max_crops_allowed(land_type, season), f"smart_greenhouse_max_{year}_{plot}_{season}"

    # 不连续重茬：同一地块同一作物不能在相邻种植季继续出现。
    crop_ids = list(data["crop_by_id"].keys())
    plant_2023 = data["plant_2023_by_plot_season"]

    for land in data["lands"]:
        plot = land["plot"]
        for cid in crop_ids:
            # 2023 对 2024 的约束。
            if cid in plant_2023.get((plot, 1), set()) or cid in plant_2023.get((plot, 2), set()):
                rid = by_plot_year_season_crop.get((plot, 2024, 1, cid))
                if rid is not None:
                    model += z[rid] == 0, f"no_replant_2023_to_2024_s1_{plot}_{cid}"
            if cid in plant_2023.get((plot, 2), set()):
                rid = by_plot_year_season_crop.get((plot, 2024, 2, cid))
                if rid is not None:
                    model += z[rid] == 0, f"no_replant_2023_s2_to_2024_s2_{plot}_{cid}"

            # 2024-2030 内部相邻季节。
            for year in YEARS:
                rid_s1 = by_plot_year_season_crop.get((plot, year, 1, cid))
                rid_s2 = by_plot_year_season_crop.get((plot, year, 2, cid))
                if rid_s1 is not None and rid_s2 is not None:
                    model += z[rid_s1] + z[rid_s2] <= 1, f"no_replant_same_year_{plot}_{year}_{cid}"
                if year < 2030:
                    next_s1 = by_plot_year_season_crop.get((plot, year + 1, 1, cid))
                    next_s2 = by_plot_year_season_crop.get((plot, year + 1, 2, cid))
                    if rid_s1 is not None and next_s1 is not None:
                        model += z[rid_s1] + z[next_s1] <= 1, f"no_replant_s1_next_s1_{plot}_{year}_{cid}"
                    if rid_s2 is not None and next_s1 is not None:
                        model += z[rid_s2] + z[next_s1] <= 1, f"no_replant_s2_next_s1_{plot}_{year}_{cid}"
                    if rid_s2 is not None and next_s2 is not None:
                        model += z[rid_s2] + z[next_s2] <= 1, f"no_replant_s2_next_s2_{plot}_{year}_{cid}"

    # 每个地块从 2023 年开始，每三年内至少种一次豆类。
    bean_ids = data["bean_ids"]
    bean_2023_flag = data["bean_2023_flag"]
    three_year_windows = [(2023, 2025), (2024, 2026), (2025, 2027), (2026, 2028), (2027, 2029), (2028, 2030)]

    for land in data["lands"]:
        plot = land["plot"]
        for start, end in three_year_windows:
            terms = []
            const = 0
            if start <= 2023 <= end:
                const += bean_2023_flag.get(plot, 0)
            for year in range(max(2024, start), min(2030, end) + 1):
                for season in SEASONS:
                    for cid in bean_ids:
                        rid = by_plot_year_season_crop.get((plot, year, season, cid))
                        if rid is not None:
                            terms.append(z[rid])
            model += pulp.lpSum(terms) + const >= 1, f"bean_{plot}_{start}_{end}"

    # 目标函数：收入 - 成本。
    # 正常销售部分 sale 按 100% 价格；超额产量按 discount_rate 价格。
    obj_terms = []
    for rid, rec in rid_to_rec.items():
        production = rec["yield"] * x[rid]
        revenue = rec["price"] * sale[rid] + discount_rate * rec["price"] * (production - sale[rid])
        cost = rec["cost"] * x[rid]
        obj_terms.append(revenue - cost)
    model += pulp.lpSum(obj_terms)

    solver = pulp.COIN_CMD(path="cbc", msg=True)
    status_code = model.solve(solver)
    status = pulp.LpStatus[status_code]

    selected: List[Dict[str, Any]] = []
    for rid, rec in rid_to_rec.items():
        area = safe_value(pulp.value(x[rid]))
        if area <= 1e-6:
            continue
        production = rec["yield"] * area
        normal_sale = min(safe_value(pulp.value(sale[rid])), production)
        excess = max(0.0, production - normal_sale)
        revenue = rec["price"] * normal_sale + discount_rate * rec["price"] * excess
        cost = rec["cost"] * area
        selected.append({
            "year": rec["year"],
            "plot": rec["plot"],
            "land_type": rec["land_type"],
            "season": rec["season"],
            "crop_id": rec["crop_id"],
            "crop_name": data["crop_name_by_id"][rec["crop_id"]],
            "area": area,
            "yield": rec["yield"],
            "production": production,
            "normal_sale": normal_sale,
            "excess": excess,
            "price": rec["price"],
            "revenue": revenue,
            "cost": cost,
            "profit": revenue - cost,
        })

    summary = {
        "case_name": case_name,
        "discount_rate": discount_rate,
        "solver_status": status,
        "objective_value": safe_value(pulp.value(model.objective)),
        "total_revenue": sum(r["revenue"] for r in selected),
        "total_cost": sum(r["cost"] for r in selected),
        "total_profit": sum(r["profit"] for r in selected),
        "total_excess_qty": sum(r["excess"] for r in selected),
        "selected_records": len(selected),
        "time_limit_seconds": time_limit,
        "gap": gap,
        "min_open_area": min_open,
        "min_greenhouse_area": min_greenhouse,
    }
    return selected, summary


def write_result_to_template(template_path: Path, output_path: Path, selected: List[Dict[str, Any]], crop_id_by_name: Dict[str, int]) -> None:
    if not template_path.exists():
        raise FileNotFoundError(f"找不到模板：{template_path}")

    wb = load_workbook(template_path)
    area_map: Dict[Tuple[int, int, str, int], float] = {}
    for row in selected:
        key = (int(row["year"]), int(row["season"]), row["plot"], int(row["crop_id"]))
        area_map[key] = area_map.get(key, 0.0) + float(row["area"])

    for year in YEARS:
        ws = wb[str(year)]

        crop_col: Dict[int, int] = {}
        for col in range(3, ws.max_column + 1):
            name = clean_text(ws.cell(row=1, column=col).value)
            if name in crop_id_by_name:
                crop_col[crop_id_by_name[name]] = col

        plot_row: Dict[Tuple[int, str], int] = {}
        for row_idx in range(2, ws.max_row + 1):
            plot = clean_text(ws.cell(row=row_idx, column=2).value)
            if not plot:
                continue
            season = 1 if row_idx <= 55 else 2
            plot_row[(season, plot)] = row_idx

        # 清空原结果区域。
        for row_idx in range(2, ws.max_row + 1):
            for col in range(3, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).value = None

        for (yy, season, plot, cid), area in area_map.items():
            if yy != year:
                continue
            r = plot_row.get((season, plot))
            c = crop_col.get(cid)
            if r is not None and c is not None and area > 1e-6:
                ws.cell(row=r, column=c).value = round(area, 4)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_summary_excel(output_path: Path, selected1: List[Dict[str, Any]], summary1: Dict[str, Any], selected2: List[Dict[str, Any]], summary2: Dict[str, Any], data: Dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "摘要"
    rows = [
        ["项目", "情形1：滞销浪费", "情形2：半价销售"],
        ["求解状态", summary1["solver_status"], summary2["solver_status"]],
        ["目标函数值", summary1["objective_value"], summary2["objective_value"]],
        ["总收入", summary1["total_revenue"], summary2["total_revenue"]],
        ["总成本", summary1["total_cost"], summary2["total_cost"]],
        ["总利润", summary1["total_profit"], summary2["total_profit"]],
        ["超出预期销售量/斤", summary1["total_excess_qty"], summary2["total_excess_qty"]],
        ["选中记录数", summary1["selected_records"], summary2["selected_records"]],
        ["MIP gap", summary1["gap"], summary2["gap"]],
        ["说明", "超过预期销售量部分不产生收入", "超过预期销售量部分按 50% 售价出售"],
    ]
    for row in rows:
        ws.append(row)
    for col in range(1, 4):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 28

    headers = ["year", "plot", "land_type", "season", "crop_id", "crop_name", "area", "yield", "production", "normal_sale", "excess", "price", "revenue", "cost", "profit"]
    for sheet_name, rows_data in [("情形1明细", selected1), ("情形2明细", selected2)]:
        wsx = wb.create_sheet(sheet_name)
        wsx.append(headers)
        for item in rows_data:
            wsx.append([item[h] for h in headers])
        for col in range(1, len(headers) + 1):
            wsx.cell(row=1, column=col).font = wsx.cell(row=1, column=col).font.copy(bold=True)
            wsx.column_dimensions[wsx.cell(row=1, column=col).column_letter].width = 15

    ws_d = wb.create_sheet("预期销售量")
    ws_d.append(["作物编号", "作物名称", "预期销售量/斤"])
    for cid, demand in sorted(data["demand"].items()):
        ws_d.append([cid, data["crop_name_by_id"][cid], demand])
    ws_d.column_dimensions["A"].width = 12
    ws_d.column_dimensions["B"].width = 14
    ws_d.column_dimensions["C"].width = 18

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="C题问题一 MILP 完整求解")
    parser.add_argument("--base", type=str, default=".", help="项目根目录，默认当前目录")
    parser.add_argument("--att1", type=str, default="附件1.xlsx", help="附件1文件名")
    parser.add_argument("--att2", type=str, default="附件2.xlsx", help="附件2文件名")
    parser.add_argument("--template-dir", type=str, default="附件3", help="模板文件夹")
    parser.add_argument("--out-dir", type=str, default="output", help="输出文件夹")
    parser.add_argument("--time-limit", type=int, default=600, help="单个情形最大求解秒数，默认 600")
    parser.add_argument("--gap", type=float, default=0.005, help="相对 MIP gap，默认 0.005")
    parser.add_argument("--min-open", type=float, default=1.0, help="露天耕地单作物最小面积，默认 1 亩")
    parser.add_argument("--min-greenhouse", type=float, default=0.1, help="大棚单作物最小面积，默认 0.1 亩")
    parser.add_argument("--quiet", action="store_true", help="不显示求解日志")
    args = parser.parse_args()

    base_dir = Path(args.base).resolve()
    template_dir = base_dir / args.template_dir
    out_dir = base_dir / args.out_dir

    print("读取附件数据...")
    data = load_data(base_dir, args.att1, args.att2)
    print(f"地块数量：{len(data['lands'])}")
    print(f"作物数量：{len(data['crops'])}")
    print(f"豆类作物数量：{len(data['bean_ids'])}")

    print("\n求解情形1：超过预期销售量部分滞销浪费...")
    selected1, summary1 = solve_one_case(
        data=data,
        discount_rate=0.0,
        case_name="waste",
        time_limit=args.time_limit,
        gap=args.gap,
        min_open=args.min_open,
        min_greenhouse=args.min_greenhouse,
        solver_msg=not args.quiet,
    )
    print(summary1)

    print("\n求解情形2：超过预期销售量部分按 50% 售价销售...")
    selected2, summary2 = solve_one_case(
        data=data,
        discount_rate=0.5,
        case_name="half_price",
        time_limit=args.time_limit,
        gap=args.gap,
        min_open=args.min_open,
        min_greenhouse=args.min_greenhouse,
        solver_msg=not args.quiet,
    )
    print(summary2)

    print("\n写回 Excel 模板...")
    write_result_to_template(template_dir / "result1_1.xlsx", out_dir / "result1_1.xlsx", selected1, data["crop_id_by_name"])
    write_result_to_template(template_dir / "result1_2.xlsx", out_dir / "result1_2.xlsx", selected2, data["crop_id_by_name"])
    write_summary_excel(out_dir / "problem1_summary.xlsx", selected1, summary1, selected2, summary2, data)

    print("\n完成。输出文件：")
    print(out_dir / "result1_1.xlsx")
    print(out_dir / "result1_2.xlsx")
    print(out_dir / "problem1_summary.xlsx")


if __name__ == "__main__":
    main()
