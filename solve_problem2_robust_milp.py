# -*- coding: utf-8 -*-
"""
2024 高教社杯数学建模 C 题：问题二鲁棒优化求解脚本

功能：
1. 读取 附件1.xlsx、附件2.xlsx；
2. 按题目给出的不确定性规则生成 2024-2030 年多情景参数；
3. 使用鲁棒混合整数线性规划（MILP）求解 2024-2030 年种植方案；
4. 写回 附件3/result2.xlsx 模板；
5. 输出 result2.xlsx、problem2_summary.xlsx、problem2_solution_long.csv。

默认目录结构：
C题/
├─ 附件1.xlsx
├─ 附件2.xlsx
├─ 附件3/
│  └─ result2.xlsx
└─ solve_problem2_robust_milp.py

运行：
python solve_problem2_robust_milp.py

指定路径运行：
python solve_problem2_robust_milp.py --base "D:\\你的路径\\C题"

重要说明：
- 这不是 baseline/贪心算法，而是带二进制决策变量的 MILP。
- 为了避免“大规模连续面积 + 二进制面积开关”导致 CBC/HiGHS 极慢，本脚本采用“单地块-单季-单作物”的整数化经营口径。
  题目允许合种，但不强制合种；该口径符合“不宜太分散、面积不宜太小”的管理要求。
- 目标函数采用保守鲁棒口径：低分位需求、低分位亩产、低分位价格与逐年增长成本。
- 如果求解达到时间限制但已有可行解，脚本仍会用当前最好可行解写出结果。
"""

from __future__ import annotations

import argparse
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.optimize import Bounds, LinearConstraint, milp
from scipy.sparse import lil_matrix

YEARS = list(range(2024, 2031))
SEASONS = [1, 2]
RICE_ID = 16
SECOND_SEASON_WATER_VEG_IDS = {35, 36, 37}
MUSHROOM_IDS = {38, 39, 40, 41}
MOREL_ID = 41
WHEAT_CORN_IDS = {6, 7}


@dataclass(frozen=True)
class Candidate:
    year: int
    plot: str
    land_type: str
    area: float
    season: int
    crop_id: int
    crop_name: str
    crop_type: str
    stat_season: str


def clean_text(x: Any) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()


def is_number(x: Any) -> bool:
    try:
        if pd.isna(x):
            return False
        float(x)
        return True
    except Exception:
        return False


def parse_price(value: Any) -> float:
    s = clean_text(value)
    if not s:
        raise ValueError("销售单价为空，无法解析")
    if "-" in s:
        a, b = s.split("-", 1)
        return (float(a) + float(b)) / 2.0
    return float(s)


def stat_season_to_model_season(s: str) -> int:
    s = clean_text(s)
    return 1 if s in ["单季", "第一季"] else 2


def load_data(base_dir: Path, att1_name: str, att2_name: str) -> Dict[str, Any]:
    att1 = base_dir / att1_name
    att2 = base_dir / att2_name
    if not att1.exists():
        raise FileNotFoundError(f"找不到 {att1}")
    if not att2.exists():
        raise FileNotFoundError(f"找不到 {att2}")

    land_df = pd.read_excel(att1, sheet_name="乡村的现有耕地")
    crop_df = pd.read_excel(att1, sheet_name="乡村种植的农作物")
    plant_df = pd.read_excel(att2, sheet_name="2023年的农作物种植情况")
    stat_df = pd.read_excel(att2, sheet_name="2023年统计的相关数据")

    lands: List[Dict[str, Any]] = []
    for _, row in land_df.iterrows():
        plot = clean_text(row.get("地块名称"))
        land_type = clean_text(row.get("地块类型"))
        area = row.get("地块面积/亩")
        if not plot or not land_type or not is_number(area):
            continue
        lands.append({"plot": plot, "land_type": land_type, "area": float(area)})
    land_by_plot = {r["plot"]: r for r in lands}

    crop_df = crop_df[pd.to_numeric(crop_df["作物编号"], errors="coerce").notna()].copy()
    crop_df["作物编号"] = crop_df["作物编号"].astype(int)
    crops: List[Dict[str, Any]] = []
    for _, row in crop_df.iterrows():
        cid = int(row["作物编号"])
        crops.append({
            "crop_id": cid,
            "crop_name": clean_text(row["作物名称"]),
            "crop_type": clean_text(row["作物类型"]),
        })
    crop_by_id = {c["crop_id"]: c for c in crops}
    crop_name_by_id = {c["crop_id"]: c["crop_name"] for c in crops}
    bean_ids = {c["crop_id"] for c in crops if "豆类" in c["crop_type"]}
    vegetable_ids = {c["crop_id"] for c in crops if "蔬菜" in c["crop_type"]}
    grain_ids = {c["crop_id"] for c in crops if "粮食" in c["crop_type"]}

    stat_df = stat_df[pd.to_numeric(stat_df["作物编号"], errors="coerce").notna()].copy()
    stat_df["作物编号"] = stat_df["作物编号"].astype(int)
    stats: Dict[Tuple[int, str, str], Dict[str, float]] = {}
    for _, row in stat_df.iterrows():
        cid = int(row["作物编号"])
        land_type = clean_text(row["地块类型"])
        stat_season = clean_text(row["种植季次"])
        stats[(cid, land_type, stat_season)] = {
            "yield": float(row["亩产量/斤"]),
            "cost": float(row["种植成本/(元/亩)"]),
            "price": parse_price(row["销售单价/(元/斤)"]),
        }

    # 题目附件说明：智慧大棚第一季蔬菜的亩产量、成本、价格与普通大棚相同，统计表省略。
    for cid in range(17, 35):
        src = (cid, "普通大棚", "第一季")
        dst = (cid, "智慧大棚", "第一季")
        if src in stats and dst not in stats:
            stats[dst] = dict(stats[src])

    plant_df["种植地块"] = plant_df["种植地块"].ffill()
    plant_df = plant_df[pd.to_numeric(plant_df["作物编号"], errors="coerce").notna()].copy()
    plant_df["作物编号"] = plant_df["作物编号"].astype(int)

    base_demand = {cid: 0.0 for cid in crop_by_id}
    plant_2023_by_plot_season: Dict[Tuple[str, int], set[int]] = {}
    bean_2023_flag = {land["plot"]: 0 for land in lands}

    for _, row in plant_df.iterrows():
        plot = clean_text(row["种植地块"])
        if plot not in land_by_plot:
            continue
        cid = int(row["作物编号"])
        area = float(row["种植面积/亩"])
        stat_season = clean_text(row["种植季次"])
        land_type = land_by_plot[plot]["land_type"]
        key = (cid, land_type, stat_season)
        if key not in stats:
            raise KeyError(f"统计参数缺失：作物 {cid}，地块类型 {land_type}，季次 {stat_season}")
        base_demand[cid] += area * stats[key]["yield"]
        season = stat_season_to_model_season(stat_season)
        plant_2023_by_plot_season.setdefault((plot, season), set()).add(cid)
        if cid in bean_ids:
            bean_2023_flag[plot] = 1

    return {
        "lands": lands,
        "land_by_plot": land_by_plot,
        "crops": crops,
        "crop_by_id": crop_by_id,
        "crop_name_by_id": crop_name_by_id,
        "bean_ids": bean_ids,
        "vegetable_ids": vegetable_ids,
        "grain_ids": grain_ids,
        "mushroom_ids": MUSHROOM_IDS,
        "stats": stats,
        "base_demand": base_demand,
        "plant_2023_by_plot_season": plant_2023_by_plot_season,
        "bean_2023_flag": bean_2023_flag,
    }


def feasible_crops_for_plot(land_type: str, season: int) -> List[Tuple[int, str]]:
    """返回某地块类型、某季允许种植的作物编号及其对应统计季次。"""
    if land_type in ["平旱地", "梯田", "山坡地"]:
        if season == 1:
            return [(cid, "单季") for cid in range(1, 16)]
        return []

    if land_type == "水浇地":
        if season == 1:
            return [(RICE_ID, "单季")] + [(cid, "第一季") for cid in range(17, 35)]
        if season == 2:
            return [(cid, "第二季") for cid in sorted(SECOND_SEASON_WATER_VEG_IDS)]
        return []

    if land_type == "普通大棚":
        if season == 1:
            return [(cid, "第一季") for cid in range(17, 35)]
        if season == 2:
            return [(cid, "第二季") for cid in range(38, 42)]
        return []

    if land_type == "智慧大棚":
        if season == 1:
            return [(cid, "第一季") for cid in range(17, 35)]
        if season == 2:
            return [(cid, "第二季") for cid in range(17, 35)]
        return []

    return []


def generate_robust_parameters(
    data: Dict[str, Any],
    n_scenarios: int,
    quantile: float,
    seed: int,
) -> Dict[str, Any]:
    """生成多情景，并取低分位参数作为鲁棒优化输入。"""
    rng = np.random.default_rng(seed)
    crops = data["crops"]
    crop_by_id = data["crop_by_id"]
    stats = data["stats"]
    base_demand = data["base_demand"]

    demand_samples: Dict[Tuple[int, int], np.ndarray] = {}
    price_samples: Dict[Tuple[int, int], np.ndarray] = {}
    yield_samples: Dict[Tuple[int, int, str, str], np.ndarray] = {}
    robust_demand: Dict[Tuple[int, int], float] = {}
    robust_price: Dict[Tuple[int, int], float] = {}
    robust_yield: Dict[Tuple[int, int, str, str], float] = {}
    mean_demand: Dict[Tuple[int, int], float] = {}
    mean_price: Dict[Tuple[int, int], float] = {}
    mean_yield: Dict[Tuple[int, int, str, str], float] = {}
    cost: Dict[Tuple[int, int, str, str], float] = {}

    # 需求情景。
    for crop in crops:
        cid = crop["crop_id"]
        base = float(base_demand.get(cid, 0.0))
        prev = np.full(n_scenarios, base, dtype=float)
        for year in YEARS:
            if cid in WHEAT_CORN_IDS:
                growth = rng.uniform(1.05, 1.10, size=n_scenarios)
                vals = prev * growth
                prev = vals
            else:
                vals = base * rng.uniform(0.95, 1.05, size=n_scenarios)
            demand_samples[(year, cid)] = vals
            robust_demand[(year, cid)] = float(np.quantile(vals, quantile))
            mean_demand[(year, cid)] = float(np.mean(vals))

    # 价格情景。粮食稳定；蔬菜年增 5%；食用菌下降，羊肚菌固定下降 5%。
    for crop in crops:
        cid = crop["crop_id"]
        crop_type = crop_by_id[cid]["crop_type"]
        base_prices = [v["price"] for (kcid, _, _), v in stats.items() if kcid == cid]
        if not base_prices:
            continue
        base_price = float(np.mean(base_prices))
        prev = np.full(n_scenarios, base_price, dtype=float)
        for year in YEARS:
            t = year - 2023
            if cid == MOREL_ID:
                vals = np.full(n_scenarios, base_price * (0.95 ** t), dtype=float)
            elif cid in MUSHROOM_IDS:
                decline = rng.uniform(0.95, 0.99, size=n_scenarios)
                vals = prev * decline
                prev = vals
            elif "蔬菜" in crop_type:
                vals = np.full(n_scenarios, base_price * (1.05 ** t), dtype=float)
            else:
                vals = np.full(n_scenarios, base_price, dtype=float)
            price_samples[(year, cid)] = vals
            robust_price[(year, cid)] = float(np.quantile(vals, quantile))
            mean_price[(year, cid)] = float(np.mean(vals))

    # 亩产量与成本。亩产量每年 ±10%；成本平均每年增长 5%。
    for (cid, land_type, stat_season), base in stats.items():
        for year in YEARS:
            vals = base["yield"] * rng.uniform(0.90, 1.10, size=n_scenarios)
            yield_samples[(year, cid, land_type, stat_season)] = vals
            robust_yield[(year, cid, land_type, stat_season)] = float(np.quantile(vals, quantile))
            mean_yield[(year, cid, land_type, stat_season)] = float(np.mean(vals))
            cost[(year, cid, land_type, stat_season)] = float(base["cost"] * (1.05 ** (year - 2023)))

    return {
        "robust_demand": robust_demand,
        "robust_price": robust_price,
        "robust_yield": robust_yield,
        "cost": cost,
        "mean_demand": mean_demand,
        "mean_price": mean_price,
        "mean_yield": mean_yield,
        "quantile": quantile,
        "n_scenarios": n_scenarios,
        "seed": seed,
    }


def build_candidates(data: Dict[str, Any]) -> List[Candidate]:
    candidates: List[Candidate] = []
    crop_by_id = data["crop_by_id"]
    stats = data["stats"]
    for year in YEARS:
        for land in data["lands"]:
            plot = land["plot"]
            land_type = land["land_type"]
            area = float(land["area"])
            for season in SEASONS:
                for cid, stat_season in feasible_crops_for_plot(land_type, season):
                    if (cid, land_type, stat_season) not in stats:
                        continue
                    crop = crop_by_id[cid]
                    candidates.append(Candidate(
                        year=year,
                        plot=plot,
                        land_type=land_type,
                        area=area,
                        season=season,
                        crop_id=cid,
                        crop_name=crop["crop_name"],
                        crop_type=crop["crop_type"],
                        stat_season=stat_season,
                    ))
    return candidates


def add_row(rows: List[Dict[int, float]], lbs: List[float], ubs: List[float], coeffs: Dict[int, float], lb: float, ub: float) -> None:
    rows.append(coeffs)
    lbs.append(lb)
    ubs.append(ub)


def solve_robust_milp(
    data: Dict[str, Any],
    params: Dict[str, Any],
    time_limit: int,
    gap: float,
    disp: bool,
) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, Any]]:
    candidates = build_candidates(data)
    n_y = len(candidates)
    crop_ids = sorted(data["crop_by_id"].keys())
    sold_keys = [(year, cid) for year in YEARS for cid in crop_ids]
    sold_index = {k: n_y + i for i, k in enumerate(sold_keys)}
    n_vars = n_y + len(sold_keys)

    cand_index: Dict[Tuple[int, str, int, int], List[int]] = {}
    cand_by_plot_year_season: Dict[Tuple[int, str, int], List[int]] = {}
    cand_by_plot_year_crop: Dict[Tuple[str, int, int], List[int]] = {}
    cand_by_plot_year_season_crop: Dict[Tuple[str, int, int, int], List[int]] = {}

    for i, c in enumerate(candidates):
        cand_index.setdefault((c.year, c.plot, c.season, c.crop_id), []).append(i)
        cand_by_plot_year_season.setdefault((c.year, c.plot, c.season), []).append(i)
        cand_by_plot_year_crop.setdefault((c.plot, c.year, c.crop_id), []).append(i)
        cand_by_plot_year_season_crop.setdefault((c.plot, c.year, c.season, c.crop_id), []).append(i)

    cvec = np.zeros(n_vars, dtype=float)
    # scipy.milp 是最小化，所以成本为正，收入为负。
    for i, cand in enumerate(candidates):
        cvec[i] = cand.area * params["cost"][(cand.year, cand.crop_id, cand.land_type, cand.stat_season)]
    for key, idx in sold_index.items():
        year, cid = key
        cvec[idx] = -params["robust_price"][(year, cid)]

    lb = np.zeros(n_vars, dtype=float)
    ub = np.ones(n_vars, dtype=float)
    ub[n_y:] = np.inf
    integrality = np.zeros(n_vars, dtype=int)
    integrality[:n_y] = 1

    rows: List[Dict[int, float]] = []
    lbs: List[float] = []
    ubs: List[float] = []

    # 1. 地块-年份-季节种植制度约束。
    for year in YEARS:
        for land in data["lands"]:
            plot = land["plot"]
            land_type = land["land_type"]
            if land_type in ["平旱地", "梯田", "山坡地"]:
                idxs = cand_by_plot_year_season.get((year, plot, 1), [])
                add_row(rows, lbs, ubs, {i: 1.0 for i in idxs}, 1.0, 1.0)

            elif land_type == "水浇地":
                s1 = cand_by_plot_year_season.get((year, plot, 1), [])
                s2 = cand_by_plot_year_season.get((year, plot, 2), [])
                add_row(rows, lbs, ubs, {i: 1.0 for i in s1}, 1.0, 1.0)
                # 第二季是否种植 = 第一季是否选择蔬菜。若第一季选择水稻，则第二季为空。
                coeffs: Dict[int, float] = {i: 1.0 for i in s2}
                for i in s1:
                    if candidates[i].crop_id != RICE_ID:
                        coeffs[i] = coeffs.get(i, 0.0) - 1.0
                add_row(rows, lbs, ubs, coeffs, 0.0, 0.0)

            elif land_type in ["普通大棚", "智慧大棚"]:
                for season in SEASONS:
                    idxs = cand_by_plot_year_season.get((year, plot, season), [])
                    add_row(rows, lbs, ubs, {i: 1.0 for i in idxs}, 1.0, 1.0)

    # 2. 同一地块连续季次不重茬。
    plant_2023 = data["plant_2023_by_plot_season"]
    for land in data["lands"]:
        plot = land["plot"]
        active_periods: List[Tuple[int, int]] = []
        # 2023 的实际记录可能只有单季，也可能两季，先放进去作为历史状态。
        for season in SEASONS:
            if (plot, season) in plant_2023:
                active_periods.append((2023, season))
        for year in YEARS:
            for season in SEASONS:
                if cand_by_plot_year_season.get((year, plot, season)):
                    active_periods.append((year, season))
        active_periods = sorted(set(active_periods))
        for (prev_y, prev_s), (cur_y, cur_s) in zip(active_periods[:-1], active_periods[1:]):
            for cid in data["crop_by_id"]:
                cur_idxs = cand_by_plot_year_season_crop.get((plot, cur_y, cur_s, cid), [])
                if not cur_idxs:
                    continue
                if prev_y == 2023:
                    if cid in plant_2023.get((plot, prev_s), set()):
                        add_row(rows, lbs, ubs, {i: 1.0 for i in cur_idxs}, 0.0, 0.0)
                else:
                    prev_idxs = cand_by_plot_year_season_crop.get((plot, prev_y, prev_s, cid), [])
                    if prev_idxs:
                        coeffs = {i: 1.0 for i in prev_idxs + cur_idxs}
                        add_row(rows, lbs, ubs, coeffs, -np.inf, 1.0)

    # 3. 三年内至少种一次豆类。按 2023-2025、2024-2026、...、2028-2030 滚动窗口。
    bean_ids = data["bean_ids"]
    bean_2023 = data["bean_2023_flag"]
    for land in data["lands"]:
        plot = land["plot"]
        for start in range(2023, 2029):
            years_in_window = [start, start + 1, start + 2]
            coeffs: Dict[int, float] = {}
            rhs_extra = 0.0
            if 2023 in years_in_window:
                rhs_extra += float(bean_2023.get(plot, 0))
            for year in years_in_window:
                if year == 2023:
                    continue
                for cid in bean_ids:
                    for i in cand_by_plot_year_crop.get((plot, year, cid), []):
                        coeffs[i] = coeffs.get(i, 0.0) + 1.0
            add_row(rows, lbs, ubs, coeffs, max(0.0, 1.0 - rhs_extra), np.inf)

    # 4. 鲁棒销量约束：sold[year,crop] <= robust_demand[year,crop]
    for year in YEARS:
        for cid in crop_ids:
            idx = sold_index[(year, cid)]
            add_row(rows, lbs, ubs, {idx: 1.0}, -np.inf, params["robust_demand"][(year, cid)])

    # 5. 鲁棒产量约束：sold[year,crop] <= sum(area * robust_yield * y)
    for year in YEARS:
        for cid in crop_ids:
            coeffs: Dict[int, float] = {sold_index[(year, cid)]: 1.0}
            for i, cand in enumerate(candidates):
                if cand.year == year and cand.crop_id == cid:
                    yld = params["robust_yield"][(year, cid, cand.land_type, cand.stat_season)]
                    coeffs[i] = coeffs.get(i, 0.0) - cand.area * yld
            add_row(rows, lbs, ubs, coeffs, -np.inf, 0.0)

    # 构造稀疏约束矩阵。
    A = lil_matrix((len(rows), n_vars), dtype=float)
    for r, coeffs in enumerate(rows):
        for j, v in coeffs.items():
            if abs(v) > 1e-12:
                A[r, j] = v
    A = A.tocsr()

    constraints = LinearConstraint(A, np.array(lbs, dtype=float), np.array(ubs, dtype=float))
    bounds = Bounds(lb, ub)

    print("========== 问题二鲁棒 MILP ==========")
    print(f"候选种植决策变量 y：{n_y}")
    print(f"销售变量 sold：{len(sold_keys)}")
    print(f"总变量数：{n_vars}")
    print(f"约束数：{len(rows)}")
    print(f"情景数：{params['n_scenarios']}，鲁棒分位数：{params['quantile']}")
    print("开始求解...")

    res = milp(
        c=cvec,
        integrality=integrality,
        bounds=bounds,
        constraints=constraints,
        options={"time_limit": time_limit, "mip_rel_gap": gap, "disp": disp},
    )

    if res.x is None:
        raise RuntimeError(f"没有得到可行解。求解状态：{res.message}")

    x = res.x
    solution_rows: List[Dict[str, Any]] = []
    for i, cand in enumerate(candidates):
        if x[i] > 0.5:
            solution_rows.append({
                "年份": cand.year,
                "地块名": cand.plot,
                "地块类型": cand.land_type,
                "季次": "第一季" if cand.season == 1 else "第二季",
                "季次编号": cand.season,
                "作物编号": cand.crop_id,
                "作物名称": cand.crop_name,
                "作物类型": cand.crop_type,
                "种植面积/亩": cand.area,
                "鲁棒亩产量/斤每亩": params["robust_yield"][(cand.year, cand.crop_id, cand.land_type, cand.stat_season)],
                "亩成本/元": params["cost"][(cand.year, cand.crop_id, cand.land_type, cand.stat_season)],
                "鲁棒售价/元每斤": params["robust_price"][(cand.year, cand.crop_id)],
                "鲁棒需求/斤": params["robust_demand"][(cand.year, cand.crop_id)],
            })
    solution_df = pd.DataFrame(solution_rows)

    sold_rows = []
    for year in YEARS:
        for cid in crop_ids:
            val = float(x[sold_index[(year, cid)]])
            if val > 1e-6:
                sold_rows.append({
                    "年份": year,
                    "作物编号": cid,
                    "作物名称": data["crop_name_by_id"][cid],
                    "鲁棒销量/斤": val,
                    "鲁棒需求/斤": params["robust_demand"][(year, cid)],
                    "鲁棒售价/元每斤": params["robust_price"][(year, cid)],
                })
    sold_df = pd.DataFrame(sold_rows)

    status_info = {
        "status": int(res.status),
        "message": str(res.message),
        "fun_minimized": float(res.fun) if res.fun is not None else None,
        "robust_profit": float(-res.fun) if res.fun is not None else None,
        "mip_gap": getattr(res, "mip_gap", None),
        "mip_node_count": getattr(res, "mip_node_count", None),
        "n_y": n_y,
        "n_vars": n_vars,
        "n_constraints": len(rows),
    }
    return solution_df, sold_df, status_info


def write_result2(template_path: Path, output_path: Path, solution_df: pd.DataFrame) -> None:
    if not template_path.exists():
        raise FileNotFoundError(f"找不到模板：{template_path}")
    wb = load_workbook(template_path)

    sol_map: Dict[Tuple[int, str, int, str], float] = {}
    for _, row in solution_df.iterrows():
        key = (int(row["年份"]), str(row["地块名"]), int(row["季次编号"]), str(row["作物名称"]))
        sol_map[key] = sol_map.get(key, 0.0) + float(row["种植面积/亩"])

    for year in YEARS:
        ws = wb[str(year)]
        crop_col: Dict[str, int] = {}
        for col in range(3, ws.max_column + 1):
            name = clean_text(ws.cell(1, col).value)
            if name:
                crop_col[name] = col

        # 清空原填报区域。
        for r in range(2, ws.max_row + 1):
            plot = clean_text(ws.cell(r, 2).value)
            if not plot or plot.startswith("("):
                continue
            for col in range(3, ws.max_column + 1):
                ws.cell(r, col).value = None

        current_season = None
        for r in range(2, ws.max_row + 1):
            marker = clean_text(ws.cell(r, 1).value)
            if "一" in marker:
                current_season = 1
            elif "二" in marker:
                current_season = 2

            plot = clean_text(ws.cell(r, 2).value)
            if not plot or current_season is None or plot.startswith("("):
                continue

            for crop_name, col in crop_col.items():
                area = sol_map.get((year, plot, current_season, crop_name), 0.0)
                if area > 1e-8:
                    ws.cell(r, col).value = round(area, 4)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_summary(data: Dict[str, Any], params: Dict[str, Any], solution_df: pd.DataFrame, sold_df: pd.DataFrame, status_info: Dict[str, Any]) -> pd.DataFrame:
    rows = []
    if solution_df.empty:
        return pd.DataFrame()

    for year in YEARS:
        plan = solution_df[solution_df["年份"] == year].copy()
        sold = sold_df[sold_df["年份"] == year].copy()
        cost_total = float((plan["种植面积/亩"] * plan["亩成本/元"]).sum())
        revenue_total = float((sold["鲁棒销量/斤"] * sold["鲁棒售价/元每斤"]).sum()) if not sold.empty else 0.0
        bean_plots = int(plan[plan["作物编号"].isin(data["bean_ids"])]["地块名"].nunique())
        rows.append({
            "年份": year,
            "种植记录数": len(plan),
            "种植面积合计/亩": float(plan["种植面积/亩"].sum()),
            "鲁棒收入/元": revenue_total,
            "成本/元": cost_total,
            "鲁棒利润/元": revenue_total - cost_total,
            "种过豆类的地块数": bean_plots,
        })
    rows.append({
        "年份": "求解状态",
        "种植记录数": status_info.get("status"),
        "种植面积合计/亩": None,
        "鲁棒收入/元": None,
        "成本/元": None,
        "鲁棒利润/元": status_info.get("robust_profit"),
        "种过豆类的地块数": status_info.get("message"),
    })
    return pd.DataFrame(rows)


def save_outputs(
    base_dir: Path,
    out_dir: Path,
    data: Dict[str, Any],
    params: Dict[str, Any],
    solution_df: pd.DataFrame,
    sold_df: pd.DataFrame,
    status_info: Dict[str, Any],
    template_name: str,
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    result_path = out_dir / "result2.xlsx"
    template_path = base_dir / "附件3" / template_name
    write_result2(template_path, result_path, solution_df)

    summary_df = build_summary(data, params, solution_df, sold_df, status_info)
    with pd.ExcelWriter(out_dir / "problem2_summary.xlsx", engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="年度摘要", index=False)
        solution_df.to_excel(writer, sheet_name="种植方案长表", index=False)
        sold_df.to_excel(writer, sheet_name="鲁棒销量", index=False)
        pd.DataFrame([status_info]).to_excel(writer, sheet_name="求解状态", index=False)
        pd.DataFrame([
            {"参数": "情景数", "取值": params["n_scenarios"]},
            {"参数": "随机种子", "取值": params["seed"]},
            {"参数": "鲁棒分位数", "取值": params["quantile"]},
            {"参数": "口径", "取值": "单地块-单季-单作物鲁棒MILP；低分位需求、低分位亩产、低分位售价、逐年增长成本"},
        ]).to_excel(writer, sheet_name="模型参数", index=False)

    solution_df.to_csv(out_dir / "problem2_solution_long.csv", index=False, encoding="utf-8-sig")
    sold_df.to_csv(out_dir / "problem2_sold.csv", index=False, encoding="utf-8-sig")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="C题问题二：鲁棒 MILP 求解脚本")
    parser.add_argument("--base", type=str, default=".", help="C题目录，默认当前目录")
    parser.add_argument("--att1", type=str, default="附件1.xlsx", help="附件1文件名")
    parser.add_argument("--att2", type=str, default="附件2.xlsx", help="附件2文件名")
    parser.add_argument("--template", type=str, default="result2.xlsx", help="附件3中的结果模板文件名")
    parser.add_argument("--out", type=str, default="output", help="输出目录")
    parser.add_argument("--scenarios", type=int, default=100, help="随机情景数量；正式跑可改成 200 或 300")
    parser.add_argument("--quantile", type=float, default=0.25, help="鲁棒低分位数，建议 0.20-0.35")
    parser.add_argument("--seed", type=int, default=20240908, help="随机种子")
    parser.add_argument("--time-limit", type=int, default=300, help="MILP 最长求解秒数")
    parser.add_argument("--gap", type=float, default=0.03, help="相对 MIP gap")
    parser.add_argument("--quiet", action="store_true", help="关闭求解器日志")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    base_dir = Path(args.base).resolve()
    out_dir = base_dir / args.out

    data = load_data(base_dir, args.att1, args.att2)
    params = generate_robust_parameters(
        data=data,
        n_scenarios=args.scenarios,
        quantile=args.quantile,
        seed=args.seed,
    )
    solution_df, sold_df, status_info = solve_robust_milp(
        data=data,
        params=params,
        time_limit=args.time_limit,
        gap=args.gap,
        disp=not args.quiet,
    )
    save_outputs(
        base_dir=base_dir,
        out_dir=out_dir,
        data=data,
        params=params,
        solution_df=solution_df,
        sold_df=sold_df,
        status_info=status_info,
        template_name=args.template,
    )

    print("\n========== 完成 ==========")
    print(f"求解状态：{status_info['message']}")
    print(f"鲁棒目标利润：{status_info['robust_profit']}")
    print(f"输出文件：{out_dir / 'result2.xlsx'}")
    print(f"摘要文件：{out_dir / 'problem2_summary.xlsx'}")
    print(f"长表文件：{out_dir / 'problem2_solution_long.csv'}")


if __name__ == "__main__":
    main()
