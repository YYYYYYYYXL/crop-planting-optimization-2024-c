# -*- coding: utf-8 -*-
"""
2024 高教社杯数学建模 C 题：问题三相关性-替代性-互补性鲁棒优化脚本

功能：
1. 读取 附件1.xlsx、附件2.xlsx；
2. 在问题二不确定性基础上，加入：
   - 作物之间的可替代性：同一替代组共享市场容量约束；
   - 作物之间的互补性：豆类轮作对后续非豆类作物给出协同收益；
   - 销售量、销售价格、种植成本之间的相关性：用相关情景模拟生成需求、价格、成本、亩产量；
3. 用相关情景的低分位参数建立鲁棒 MILP；
4. 输出 result3.xlsx、problem3_summary.xlsx、problem3_solution_long.csv；
5. 若 output/problem2_solution_long.csv 存在，会自动用同一批问题三情景评估问题二方案，并输出比较表。

默认目录结构：
C题/
├─ 附件1.xlsx
├─ 附件2.xlsx
├─ 附件3/
│  └─ result2.xlsx
└─ solve_problem3_correlated_robust.py

运行：
python solve_problem3_correlated_robust.py

指定路径运行：
python solve_problem3_correlated_robust.py --base "D:\\你的路径\\C题"

说明：
- 问题三没有官方唯一结果模板，本脚本使用附件3/result2.xlsx 的同结构模板写出 output/result3.xlsx。
- 这不是简单 baseline：模型显式加入了相关情景模拟、替代组市场容量、豆类轮作互补收益和多情景风险评估。
- 为控制求解时间，经营口径沿用“单地块-单季-单作物”。题目允许合种但不强制合种，该口径也符合“不宜太分散、面积不宜太小”。
"""

from __future__ import annotations

import argparse
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.optimize import Bounds, LinearConstraint, milp
from scipy.sparse import lil_matrix

YEARS = list(range(2024, 2031))
SEASONS = [1, 2]
RICE_ID = 16
SECOND_SEASON_WATER_VEG_IDS = {35, 36, 37}
MUSHROOM_IDS = {38, 39, 40, 41}
MOREL_ID = 41
WHEAT_CORN_IDS = {6, 7}


@dataclass(frozen=True)
class Candidate:
    year: int
    plot: str
    land_type: str
    area: float
    season: int
    crop_id: int
    crop_name: str
    crop_type: str
    stat_season: str


def clean_text(x: Any) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()


def is_number(x: Any) -> bool:
    try:
        if pd.isna(x):
            return False
        float(x)
        return True
    except Exception:
        return False


def parse_price(value: Any) -> float:
    s = clean_text(value)
    if not s:
        raise ValueError("销售单价为空，无法解析")
    if "-" in s:
        a, b = s.split("-", 1)
        return (float(a) + float(b)) / 2.0
    return float(s)


def stat_season_to_model_season(s: str) -> int:
    s = clean_text(s)
    return 1 if s in ["单季", "第一季"] else 2


def load_data(base_dir: Path, att1_name: str, att2_name: str) -> Dict[str, Any]:
    att1 = base_dir / att1_name
    att2 = base_dir / att2_name
    if not att1.exists():
        raise FileNotFoundError(f"找不到 {att1}")
    if not att2.exists():
        raise FileNotFoundError(f"找不到 {att2}")

    land_df = pd.read_excel(att1, sheet_name="乡村的现有耕地")
    crop_df = pd.read_excel(att1, sheet_name="乡村种植的农作物")
    plant_df = pd.read_excel(att2, sheet_name="2023年的农作物种植情况")
    stat_df = pd.read_excel(att2, sheet_name="2023年统计的相关数据")

    lands: List[Dict[str, Any]] = []
    for _, row in land_df.iterrows():
        plot = clean_text(row.get("地块名称"))
        land_type = clean_text(row.get("地块类型"))
        area = row.get("地块面积/亩")
        if not plot or not land_type or not is_number(area):
            continue
        lands.append({"plot": plot, "land_type": land_type, "area": float(area)})
    land_by_plot = {r["plot"]: r for r in lands}

    crop_df = crop_df[pd.to_numeric(crop_df["作物编号"], errors="coerce").notna()].copy()
    crop_df["作物编号"] = crop_df["作物编号"].astype(int)
    crops: List[Dict[str, Any]] = []
    for _, row in crop_df.iterrows():
        cid = int(row["作物编号"])
        crops.append({
            "crop_id": cid,
            "crop_name": clean_text(row["作物名称"]),
            "crop_type": clean_text(row["作物类型"]),
        })
    crop_by_id = {c["crop_id"]: c for c in crops}
    crop_name_by_id = {c["crop_id"]: c["crop_name"] for c in crops}
    bean_ids = {c["crop_id"] for c in crops if "豆类" in c["crop_type"]}
    vegetable_ids = {c["crop_id"] for c in crops if "蔬菜" in c["crop_type"]}
    grain_ids = {c["crop_id"] for c in crops if "粮食" in c["crop_type"]}

    stat_df = stat_df[pd.to_numeric(stat_df["作物编号"], errors="coerce").notna()].copy()
    stat_df["作物编号"] = stat_df["作物编号"].astype(int)
    stats: Dict[Tuple[int, str, str], Dict[str, float]] = {}
    for _, row in stat_df.iterrows():
        cid = int(row["作物编号"])
        land_type = clean_text(row["地块类型"])
        stat_season = clean_text(row["种植季次"])
        stats[(cid, land_type, stat_season)] = {
            "yield": float(row["亩产量/斤"]),
            "cost": float(row["种植成本/(元/亩)"]),
            "price": parse_price(row["销售单价/(元/斤)"]),
        }

    # 附件说明：智慧大棚第一季蔬菜参数与普通大棚第一季相同，原表省略。
    for cid in range(17, 35):
        src = (cid, "普通大棚", "第一季")
        dst = (cid, "智慧大棚", "第一季")
        if src in stats and dst not in stats:
            stats[dst] = dict(stats[src])

    plant_df["种植地块"] = plant_df["种植地块"].ffill()
    plant_df = plant_df[pd.to_numeric(plant_df["作物编号"], errors="coerce").notna()].copy()
    plant_df["作物编号"] = plant_df["作物编号"].astype(int)

    base_demand = {cid: 0.0 for cid in crop_by_id}
    plant_2023_by_plot_season: Dict[Tuple[str, int], set[int]] = {}
    bean_2023_flag = {land["plot"]: 0 for land in lands}

    for _, row in plant_df.iterrows():
        plot = clean_text(row["种植地块"])
        if plot not in land_by_plot:
            continue
        cid = int(row["作物编号"])
        area = float(row["种植面积/亩"])
        stat_season = clean_text(row["种植季次"])
        land_type = land_by_plot[plot]["land_type"]
        key = (cid, land_type, stat_season)
        if key not in stats:
            raise KeyError(f"统计参数缺失：作物 {cid}，地块类型 {land_type}，季次 {stat_season}")
        base_demand[cid] += area * stats[key]["yield"]
        season = stat_season_to_model_season(stat_season)
        plant_2023_by_plot_season.setdefault((plot, season), set()).add(cid)
        if cid in bean_ids:
            bean_2023_flag[plot] = 1

    return {
        "lands": lands,
        "land_by_plot": land_by_plot,
        "crops": crops,
        "crop_by_id": crop_by_id,
        "crop_name_by_id": crop_name_by_id,
        "bean_ids": bean_ids,
        "vegetable_ids": vegetable_ids,
        "grain_ids": grain_ids,
        "mushroom_ids": MUSHROOM_IDS,
        "stats": stats,
        "base_demand": base_demand,
        "plant_2023_by_plot_season": plant_2023_by_plot_season,
        "bean_2023_flag": bean_2023_flag,
    }


def feasible_crops_for_plot(land_type: str, season: int) -> List[Tuple[int, str]]:
    if land_type in ["平旱地", "梯田", "山坡地"]:
        if season == 1:
            return [(cid, "单季") for cid in range(1, 16)]
        return []

    if land_type == "水浇地":
        if season == 1:
            return [(RICE_ID, "单季")] + [(cid, "第一季") for cid in range(17, 35)]
        if season == 2:
            return [(cid, "第二季") for cid in sorted(SECOND_SEASON_WATER_VEG_IDS)]
        return []

    if land_type == "普通大棚":
        if season == 1:
            return [(cid, "第一季") for cid in range(17, 35)]
        if season == 2:
            return [(cid, "第二季") for cid in range(38, 42)]
        return []

    if land_type == "智慧大棚":
        if season == 1:
            return [(cid, "第一季") for cid in range(17, 35)]
        if season == 2:
            return [(cid, "第二季") for cid in range(17, 35)]
        return []

    return []


def crop_substitution_group(cid: int) -> str:
    """问题三替代组。组内作物视为市场替代品，共享一部分市场容量。"""
    if cid in {1, 2, 3, 4, 5}:
        return "豆类粮食"
    if cid in {6, 7, 16}:
        return "主粮"
    if cid in {8, 9, 10, 11, 12, 13, 14, 15}:
        return "杂粮薯类"
    if 17 <= cid <= 34:
        return "第一季蔬菜"
    if cid in {35, 36, 37}:
        return "第二季根茎类蔬菜"
    if cid in MUSHROOM_IDS:
        return "食用菌"
    return "其他"


def build_substitution_groups(crop_ids: List[int]) -> Dict[str, List[int]]:
    groups: Dict[str, List[int]] = {}
    for cid in crop_ids:
        groups.setdefault(crop_substitution_group(cid), []).append(cid)
    return groups


def build_corr_matrix(group_names: List[str]) -> np.ndarray:
    """
    为替代组需求冲击构造相关矩阵。
    组内替代由组共同因子体现；组间保留中等正相关，主粮和蔬菜相关性稍低。
    """
    n = len(group_names)
    corr = np.full((n, n), 0.25, dtype=float)
    np.fill_diagonal(corr, 1.0)
    for i, g1 in enumerate(group_names):
        for j, g2 in enumerate(group_names):
            if i == j:
                continue
            if "蔬菜" in g1 and "蔬菜" in g2:
                corr[i, j] = 0.55
            elif ("主粮" in g1 and "杂粮" in g2) or ("杂粮" in g1 and "主粮" in g2):
                corr[i, j] = 0.50
            elif "食用菌" in g1 or "食用菌" in g2:
                corr[i, j] = 0.18
            elif "豆类" in g1 or "豆类" in g2:
                corr[i, j] = 0.35
    # 数值修正，保证半正定。
    vals, vecs = np.linalg.eigh(corr)
    vals = np.maximum(vals, 1e-6)
    corr_psd = (vecs @ np.diag(vals) @ vecs.T)
    d = np.sqrt(np.diag(corr_psd))
    return corr_psd / np.outer(d, d)


def generate_correlated_parameters(
    data: Dict[str, Any],
    n_scenarios: int,
    quantile: float,
    seed: int,
    group_capacity_alpha: float,
) -> Dict[str, Any]:
    """
    相关情景模拟。
    - 需求冲击按替代组相关；
    - 价格与需求冲击部分负相关；
    - 成本与年度通胀冲击正相关；
    - 亩产量与天气冲击相关。
    """
    rng = np.random.default_rng(seed)
    crops = data["crops"]
    crop_by_id = data["crop_by_id"]
    stats = data["stats"]
    base_demand = data["base_demand"]
    crop_ids = sorted(data["crop_by_id"].keys())
    groups = build_substitution_groups(crop_ids)
    group_names = sorted(groups.keys())
    group_index = {g: i for i, g in enumerate(group_names)}
    corr = build_corr_matrix(group_names)

    # 样本容器。
    demand_samples: Dict[Tuple[int, int], np.ndarray] = {}
    price_samples: Dict[Tuple[int, int], np.ndarray] = {}
    yield_samples: Dict[Tuple[int, int, str, str], np.ndarray] = {}
    cost_samples: Dict[Tuple[int, int, str, str], np.ndarray] = {}
    group_capacity_samples: Dict[Tuple[int, str], np.ndarray] = {}

    robust_demand: Dict[Tuple[int, int], float] = {}
    robust_price: Dict[Tuple[int, int], float] = {}
    robust_yield: Dict[Tuple[int, int, str, str], float] = {}
    robust_cost: Dict[Tuple[int, int, str, str], float] = {}
    robust_group_capacity: Dict[Tuple[int, str], float] = {}

    mean_demand: Dict[Tuple[int, int], float] = {}
    mean_price: Dict[Tuple[int, int], float] = {}
    mean_yield: Dict[Tuple[int, int, str, str], float] = {}
    mean_cost: Dict[Tuple[int, int, str, str], float] = {}

    # 每年公共因子。
    prev_wheat_corn = {cid: np.full(n_scenarios, float(base_demand.get(cid, 0.0))) for cid in WHEAT_CORN_IDS}
    prev_mushroom_price: Dict[int, np.ndarray] = {}

    for year in YEARS:
        t = year - 2023
        group_shock = rng.multivariate_normal(np.zeros(len(group_names)), corr, size=n_scenarios)
        global_demand = rng.normal(0, 1, size=n_scenarios)
        inflation = rng.normal(0, 1, size=n_scenarios)
        weather = rng.normal(0, 1, size=n_scenarios)

        # 需求：严格保留题目给出的主要区间，只是通过共同因子引入相关性。
        for crop in crops:
            cid = crop["crop_id"]
            group = crop_substitution_group(cid)
            gi = group_index[group]
            z = 0.70 * group_shock[:, gi] + 0.30 * global_demand + rng.normal(0, 0.35, size=n_scenarios)
            bounded = np.tanh(z)
            base = float(base_demand.get(cid, 0.0))
            if cid in WHEAT_CORN_IDS:
                growth = 1.075 + 0.025 * bounded  # [1.05, 1.10]
                vals = prev_wheat_corn[cid] * growth
                prev_wheat_corn[cid] = vals
            else:
                vals = base * (1.0 + 0.05 * bounded)  # [0.95, 1.05]
            demand_samples[(year, cid)] = vals
            robust_demand[(year, cid)] = float(np.quantile(vals, quantile))
            mean_demand[(year, cid)] = float(np.mean(vals))

        # 替代组共享市场容量：组内作物不是完全独立需求，合计销量需受组市场容量约束。
        for group, cids in groups.items():
            total = np.zeros(n_scenarios, dtype=float)
            for cid in cids:
                total += demand_samples[(year, cid)]
            # alpha 越小，替代关系越强，组总容量越紧。
            cap = group_capacity_alpha * total
            group_capacity_samples[(year, group)] = cap
            robust_group_capacity[(year, group)] = float(np.quantile(cap, quantile))

        # 价格：与需求冲击弱负相关，与通胀弱正相关。
        for crop in crops:
            cid = crop["crop_id"]
            crop_type = crop_by_id[cid]["crop_type"]
            group = crop_substitution_group(cid)
            gi = group_index[group]
            z = -0.45 * group_shock[:, gi] + 0.35 * inflation + rng.normal(0, 0.25, size=n_scenarios)
            bounded = np.tanh(z)
            base_prices = [v["price"] for (kcid, _, _), v in stats.items() if kcid == cid]
            if not base_prices:
                continue
            base_price = float(np.mean(base_prices))
            if cid == MOREL_ID:
                vals = base_price * (0.95 ** t) * (1.0 + 0.01 * bounded)
            elif cid in MUSHROOM_IDS:
                if cid not in prev_mushroom_price:
                    prev_mushroom_price[cid] = np.full(n_scenarios, base_price, dtype=float)
                decline = 0.97 + 0.02 * bounded  # 约 [0.95, 0.99]
                vals = prev_mushroom_price[cid] * decline
                prev_mushroom_price[cid] = vals
            elif "蔬菜" in crop_type:
                vals = base_price * (1.05 ** t) * (1.0 + 0.03 * bounded)
            else:
                vals = base_price * (1.0 + 0.02 * bounded)
            vals = np.maximum(vals, 1e-6)
            price_samples[(year, cid)] = vals
            robust_price[(year, cid)] = float(np.quantile(vals, quantile))
            mean_price[(year, cid)] = float(np.mean(vals))

        # 亩产量和成本：天气影响产量；成本随 5% 年增长并受通胀正相关影响。
        for (cid, land_type, stat_season), base in stats.items():
            group = crop_substitution_group(cid)
            gi = group_index[group]
            yz = 0.65 * weather + 0.25 * group_shock[:, gi] + rng.normal(0, 0.35, size=n_scenarios)
            yvals = base["yield"] * (1.0 + 0.10 * np.tanh(yz))
            cost_z = 0.70 * inflation + 0.20 * group_shock[:, gi] + rng.normal(0, 0.25, size=n_scenarios)
            # 成本以年均 5% 增长为中心，上下小幅波动，避免偏离题意过大。
            cvals = base["cost"] * (1.05 ** t) * (1.0 + 0.03 * np.tanh(cost_z))
            cvals = np.maximum(cvals, 1e-6)
            yield_samples[(year, cid, land_type, stat_season)] = yvals
            cost_samples[(year, cid, land_type, stat_season)] = cvals
            robust_yield[(year, cid, land_type, stat_season)] = float(np.quantile(yvals, quantile))
            # 成本对利润不利，取高分位更保守。
            robust_cost[(year, cid, land_type, stat_season)] = float(np.quantile(cvals, 1.0 - quantile))
            mean_yield[(year, cid, land_type, stat_season)] = float(np.mean(yvals))
            mean_cost[(year, cid, land_type, stat_season)] = float(np.mean(cvals))

    return {
        "demand_samples": demand_samples,
        "price_samples": price_samples,
        "yield_samples": yield_samples,
        "cost_samples": cost_samples,
        "group_capacity_samples": group_capacity_samples,
        "robust_demand": robust_demand,
        "robust_price": robust_price,
        "robust_yield": robust_yield,
        "robust_cost": robust_cost,
        "robust_group_capacity": robust_group_capacity,
        "mean_demand": mean_demand,
        "mean_price": mean_price,
        "mean_yield": mean_yield,
        "mean_cost": mean_cost,
        "groups": groups,
        "group_names": group_names,
        "group_corr": corr,
        "quantile": quantile,
        "n_scenarios": n_scenarios,
        "seed": seed,
        "group_capacity_alpha": group_capacity_alpha,
    }


def build_candidates(data: Dict[str, Any]) -> List[Candidate]:
    candidates: List[Candidate] = []
    crop_by_id = data["crop_by_id"]
    stats = data["stats"]
    for year in YEARS:
        for land in data["lands"]:
            plot = land["plot"]
            land_type = land["land_type"]
            area = float(land["area"])
            for season in SEASONS:
                for cid, stat_season in feasible_crops_for_plot(land_type, season):
                    if (cid, land_type, stat_season) not in stats:
                        continue
                    crop = crop_by_id[cid]
                    candidates.append(Candidate(
                        year=year,
                        plot=plot,
                        land_type=land_type,
                        area=area,
                        season=season,
                        crop_id=cid,
                        crop_name=crop["crop_name"],
                        crop_type=crop["crop_type"],
                        stat_season=stat_season,
                    ))
    return candidates


def add_row(rows: List[Dict[int, float]], lbs: List[float], ubs: List[float], coeffs: Dict[int, float], lb: float, ub: float) -> None:
    rows.append(coeffs)
    lbs.append(lb)
    ubs.append(ub)


def solve_problem3_milp(
    data: Dict[str, Any],
    params: Dict[str, Any],
    complement_bonus: float,
    time_limit: int,
    gap: float,
    disp: bool,
) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, Any]]:
    candidates = build_candidates(data)
    n_y = len(candidates)
    crop_ids = sorted(data["crop_by_id"].keys())
    sold_keys = [(year, cid) for year in YEARS for cid in crop_ids]
    sold_index = {k: n_y + i for i, k in enumerate(sold_keys)}

    # 互补性变量：B/N/C。
    # B[p,y]=该地块该年是否种豆类；N[p,y]=该地块该年是否种非豆类；C[p,y]=上一年豆类 + 当年非豆类的互补轮作。
    plot_year_keys = [(land["plot"], year) for land in data["lands"] for year in YEARS]
    b_start = n_y + len(sold_keys)
    n_start = b_start + len(plot_year_keys)
    c_start = n_start + len(plot_year_keys)
    B_index = {k: b_start + i for i, k in enumerate(plot_year_keys)}
    N_index = {k: n_start + i for i, k in enumerate(plot_year_keys)}
    C_index = {k: c_start + i for i, k in enumerate(plot_year_keys)}
    n_vars = c_start + len(plot_year_keys)

    cand_by_plot_year_season: Dict[Tuple[int, str, int], List[int]] = {}
    cand_by_plot_year_crop: Dict[Tuple[str, int, int], List[int]] = {}
    cand_by_plot_year_season_crop: Dict[Tuple[str, int, int, int], List[int]] = {}
    cand_by_plot_year: Dict[Tuple[str, int], List[int]] = {}
    cand_by_plot_year_bean: Dict[Tuple[str, int], List[int]] = {}
    cand_by_plot_year_nonbean: Dict[Tuple[str, int], List[int]] = {}

    bean_ids = data["bean_ids"]
    for i, cand in enumerate(candidates):
        cand_by_plot_year_season.setdefault((cand.year, cand.plot, cand.season), []).append(i)
        cand_by_plot_year_crop.setdefault((cand.plot, cand.year, cand.crop_id), []).append(i)
        cand_by_plot_year_season_crop.setdefault((cand.plot, cand.year, cand.season, cand.crop_id), []).append(i)
        cand_by_plot_year.setdefault((cand.plot, cand.year), []).append(i)
        if cand.crop_id in bean_ids:
            cand_by_plot_year_bean.setdefault((cand.plot, cand.year), []).append(i)
        else:
            cand_by_plot_year_nonbean.setdefault((cand.plot, cand.year), []).append(i)

    cvec = np.zeros(n_vars, dtype=float)
    # y 变量承担种植成本；sold 变量获得销售收入。
    for i, cand in enumerate(candidates):
        cvec[i] = cand.area * params["robust_cost"][(cand.year, cand.crop_id, cand.land_type, cand.stat_season)]
    for key, idx in sold_index.items():
        year, cid = key
        cvec[idx] = -params["robust_price"][(year, cid)]

    # 互补收益：上一年种豆类、当年种非豆类，视为土壤养分/轮作协同带来的稳健收益。
    # 该参数是论文中的敏感性参数，默认 3%。
    avg_margin_by_year: Dict[int, float] = {}
    for year in YEARS:
        margins = []
        for cand in candidates:
            if cand.year != year:
                continue
            yld = params["robust_yield"][(year, cand.crop_id, cand.land_type, cand.stat_season)]
            price = params["robust_price"][(year, cand.crop_id)]
            cost = params["robust_cost"][(year, cand.crop_id, cand.land_type, cand.stat_season)]
            margins.append(max(0.0, yld * price - cost))
        avg_margin_by_year[year] = float(np.mean(margins)) if margins else 0.0

    for land in data["lands"]:
        plot = land["plot"]
        area = float(land["area"])
        for year in YEARS:
            # scipy.milp 最小化，所以收益为负号。
            cvec[C_index[(plot, year)]] = -complement_bonus * area * avg_margin_by_year[year]

    lb = np.zeros(n_vars, dtype=float)
    ub = np.ones(n_vars, dtype=float)
    ub[n_y:n_y + len(sold_keys)] = np.inf
    integrality = np.zeros(n_vars, dtype=int)
    integrality[:n_y] = 1
    integrality[b_start:] = 1

    rows: List[Dict[int, float]] = []
    lbs: List[float] = []
    ubs: List[float] = []

    # 1. 地块-年份-季节种植制度。
    for year in YEARS:
        for land in data["lands"]:
            plot = land["plot"]
            land_type = land["land_type"]
            if land_type in ["平旱地", "梯田", "山坡地"]:
                idxs = cand_by_plot_year_season.get((year, plot, 1), [])
                add_row(rows, lbs, ubs, {i: 1.0 for i in idxs}, 1.0, 1.0)

            elif land_type == "水浇地":
                s1 = cand_by_plot_year_season.get((year, plot, 1), [])
                s2 = cand_by_plot_year_season.get((year, plot, 2), [])
                add_row(rows, lbs, ubs, {i: 1.0 for i in s1}, 1.0, 1.0)
                coeffs: Dict[int, float] = {i: 1.0 for i in s2}
                for i in s1:
                    if candidates[i].crop_id != RICE_ID:
                        coeffs[i] = coeffs.get(i, 0.0) - 1.0
                add_row(rows, lbs, ubs, coeffs, 0.0, 0.0)

            elif land_type in ["普通大棚", "智慧大棚"]:
                for season in SEASONS:
                    idxs = cand_by_plot_year_season.get((year, plot, season), [])
                    add_row(rows, lbs, ubs, {i: 1.0 for i in idxs}, 1.0, 1.0)

    # 2. 连续季次不重茬。
    plant_2023 = data["plant_2023_by_plot_season"]
    for land in data["lands"]:
        plot = land["plot"]
        active_periods: List[Tuple[int, int]] = []
        for season in SEASONS:
            if (plot, season) in plant_2023:
                active_periods.append((2023, season))
        for year in YEARS:
            for season in SEASONS:
                if cand_by_plot_year_season.get((year, plot, season)):
                    active_periods.append((year, season))
        active_periods = sorted(set(active_periods))
        for (prev_y, prev_s), (cur_y, cur_s) in zip(active_periods[:-1], active_periods[1:]):
            for cid in data["crop_by_id"]:
                cur_idxs = cand_by_plot_year_season_crop.get((plot, cur_y, cur_s, cid), [])
                if not cur_idxs:
                    continue
                if prev_y == 2023:
                    if cid in plant_2023.get((plot, prev_s), set()):
                        add_row(rows, lbs, ubs, {i: 1.0 for i in cur_idxs}, 0.0, 0.0)
                else:
                    prev_idxs = cand_by_plot_year_season_crop.get((plot, prev_y, prev_s, cid), [])
                    if prev_idxs:
                        coeffs = {i: 1.0 for i in prev_idxs + cur_idxs}
                        add_row(rows, lbs, ubs, coeffs, -np.inf, 1.0)

    # 3. 三年内至少种一次豆类。
    bean_2023 = data["bean_2023_flag"]
    for land in data["lands"]:
        plot = land["plot"]
        for start in range(2023, 2029):
            years_in_window = [start, start + 1, start + 2]
            coeffs: Dict[int, float] = {}
            rhs_extra = 0.0
            if 2023 in years_in_window:
                rhs_extra += float(bean_2023.get(plot, 0))
            for year in years_in_window:
                if year == 2023:
                    continue
                for i in cand_by_plot_year_bean.get((plot, year), []):
                    coeffs[i] = coeffs.get(i, 0.0) + 1.0
            add_row(rows, lbs, ubs, coeffs, max(0.0, 1.0 - rhs_extra), np.inf)

    # 4. 单作物销量约束：销量 <= 鲁棒需求。
    for year in YEARS:
        for cid in crop_ids:
            idx = sold_index[(year, cid)]
            add_row(rows, lbs, ubs, {idx: 1.0}, -np.inf, params["robust_demand"][(year, cid)])

    # 5. 单作物产量约束：销量 <= 鲁棒产量。
    for year in YEARS:
        for cid in crop_ids:
            coeffs: Dict[int, float] = {sold_index[(year, cid)]: 1.0}
            for i, cand in enumerate(candidates):
                if cand.year == year and cand.crop_id == cid:
                    yld = params["robust_yield"][(year, cid, cand.land_type, cand.stat_season)]
                    coeffs[i] = coeffs.get(i, 0.0) - cand.area * yld
            add_row(rows, lbs, ubs, coeffs, -np.inf, 0.0)

    # 6. 替代组市场容量约束：组内销量合计 <= 组鲁棒市场容量。
    for year in YEARS:
        for group, cids in params["groups"].items():
            coeffs = {sold_index[(year, cid)]: 1.0 for cid in cids}
            add_row(rows, lbs, ubs, coeffs, -np.inf, params["robust_group_capacity"][(year, group)])

    # 7. B/N/C 互补性线性化。
    for land in data["lands"]:
        plot = land["plot"]
        for year in YEARS:
            all_idxs = cand_by_plot_year.get((plot, year), [])
            bean_idxs = cand_by_plot_year_bean.get((plot, year), [])
            nonbean_idxs = cand_by_plot_year_nonbean.get((plot, year), [])
            M_all = max(1, len(all_idxs))
            M_bean = max(1, len(bean_idxs))
            M_nonbean = max(1, len(nonbean_idxs))
            bidx = B_index[(plot, year)]
            nidx = N_index[(plot, year)]
            cidx = C_index[(plot, year)]

            # B = 是否有任意豆类候选被选中。
            coeffs = {i: 1.0 for i in bean_idxs}
            coeffs[bidx] = coeffs.get(bidx, 0.0) - 1.0
            add_row(rows, lbs, ubs, coeffs, 0.0, np.inf)  # sum_bean >= B
            coeffs = {i: 1.0 for i in bean_idxs}
            coeffs[bidx] = coeffs.get(bidx, 0.0) - float(M_bean)
            add_row(rows, lbs, ubs, coeffs, -np.inf, 0.0)  # sum_bean <= M*B

            # N = 是否有任意非豆类候选被选中。
            coeffs = {i: 1.0 for i in nonbean_idxs}
            coeffs[nidx] = coeffs.get(nidx, 0.0) - 1.0
            add_row(rows, lbs, ubs, coeffs, 0.0, np.inf)
            coeffs = {i: 1.0 for i in nonbean_idxs}
            coeffs[nidx] = coeffs.get(nidx, 0.0) - float(M_nonbean)
            add_row(rows, lbs, ubs, coeffs, -np.inf, 0.0)

            # C = 上一年 B 与当年 N 的交集；2024 使用 2023 历史豆类标记。
            if year == 2024:
                prev_const = float(bean_2023.get(plot, 0))
                if prev_const >= 0.5:
                    # C = N
                    add_row(rows, lbs, ubs, {cidx: 1.0, nidx: -1.0}, 0.0, 0.0)
                else:
                    # C = 0
                    add_row(rows, lbs, ubs, {cidx: 1.0}, 0.0, 0.0)
            else:
                prev_bidx = B_index[(plot, year - 1)]
                add_row(rows, lbs, ubs, {cidx: 1.0, prev_bidx: -1.0}, -np.inf, 0.0)  # C <= Bprev
                add_row(rows, lbs, ubs, {cidx: 1.0, nidx: -1.0}, -np.inf, 0.0)       # C <= N
                add_row(rows, lbs, ubs, {cidx: 1.0, prev_bidx: -1.0, nidx: -1.0}, -1.0, np.inf)  # C >= Bprev + N - 1

    A = lil_matrix((len(rows), n_vars), dtype=float)
    for r, coeffs in enumerate(rows):
        for j, v in coeffs.items():
            if abs(v) > 1e-12:
                A[r, j] = v
    A = A.tocsr()

    constraints = LinearConstraint(A, np.array(lbs, dtype=float), np.array(ubs, dtype=float))
    bounds = Bounds(lb, ub)

    print("========== 问题三：相关性-替代性-互补性鲁棒 MILP ==========")
    print(f"候选种植决策变量 y：{n_y}")
    print(f"销售变量 sold：{len(sold_keys)}")
    print(f"互补性变量 B/N/C：{3 * len(plot_year_keys)}")
    print(f"总变量数：{n_vars}")
    print(f"约束数：{len(rows)}")
    print(f"情景数：{params['n_scenarios']}，鲁棒分位数：{params['quantile']}")
    print(f"替代组容量系数：{params['group_capacity_alpha']}，互补收益系数：{complement_bonus}")
    print("开始求解...")

    res = milp(
        c=cvec,
        integrality=integrality,
        bounds=bounds,
        constraints=constraints,
        options={"time_limit": time_limit, "mip_rel_gap": gap, "disp": disp},
    )

    if res.x is None:
        raise RuntimeError(f"没有得到可行解。求解状态：{res.message}")

    x = res.x
    solution_rows: List[Dict[str, Any]] = []
    for i, cand in enumerate(candidates):
        if x[i] > 0.5:
            solution_rows.append({
                "年份": cand.year,
                "地块名": cand.plot,
                "地块类型": cand.land_type,
                "季次": "第一季" if cand.season == 1 else "第二季",
                "季次编号": cand.season,
                "作物编号": cand.crop_id,
                "作物名称": cand.crop_name,
                "作物类型": cand.crop_type,
                "替代组": crop_substitution_group(cand.crop_id),
                "统计季次": cand.stat_season,
                "种植面积/亩": cand.area,
                "鲁棒亩产量/斤每亩": params["robust_yield"][(cand.year, cand.crop_id, cand.land_type, cand.stat_season)],
                "亩成本/元": params["robust_cost"][(cand.year, cand.crop_id, cand.land_type, cand.stat_season)],
                "鲁棒售价/元每斤": params["robust_price"][(cand.year, cand.crop_id)],
                "鲁棒需求/斤": params["robust_demand"][(cand.year, cand.crop_id)],
            })
    solution_df = pd.DataFrame(solution_rows)

    sold_rows = []
    for year in YEARS:
        for cid in crop_ids:
            val = float(x[sold_index[(year, cid)]])
            if val > 1e-6:
                sold_rows.append({
                    "年份": year,
                    "作物编号": cid,
                    "作物名称": data["crop_name_by_id"][cid],
                    "替代组": crop_substitution_group(cid),
                    "鲁棒销量/斤": val,
                    "鲁棒需求/斤": params["robust_demand"][(year, cid)],
                    "鲁棒售价/元每斤": params["robust_price"][(year, cid)],
                })
    sold_df = pd.DataFrame(sold_rows)

    status_info = {
        "status": int(res.status),
        "message": str(res.message),
        "fun_minimized": float(res.fun) if res.fun is not None else None,
        "robust_objective_with_complement": float(-res.fun) if res.fun is not None else None,
        "mip_gap": getattr(res, "mip_gap", None),
        "mip_node_count": getattr(res, "mip_node_count", None),
        "n_y": n_y,
        "n_vars": n_vars,
        "n_constraints": len(rows),
    }
    return solution_df, sold_df, status_info


def write_result_template(template_path: Path, output_path: Path, solution_df: pd.DataFrame) -> None:
    if not template_path.exists():
        raise FileNotFoundError(f"找不到模板：{template_path}")
    wb = load_workbook(template_path)

    sol_map: Dict[Tuple[int, str, int, str], float] = {}
    for _, row in solution_df.iterrows():
        key = (int(row["年份"]), str(row["地块名"]), int(row["季次编号"]), str(row["作物名称"]))
        sol_map[key] = sol_map.get(key, 0.0) + float(row["种植面积/亩"])

    for year in YEARS:
        ws = wb[str(year)]
        crop_col: Dict[str, int] = {}
        for col in range(3, ws.max_column + 1):
            name = clean_text(ws.cell(1, col).value)
            if name:
                crop_col[name] = col

        for r in range(2, ws.max_row + 1):
            plot = clean_text(ws.cell(r, 2).value)
            if not plot or plot.startswith("("):
                continue
            for col in range(3, ws.max_column + 1):
                ws.cell(r, col).value = None

        current_season = None
        for r in range(2, ws.max_row + 1):
            marker = clean_text(ws.cell(r, 1).value)
            if "一" in marker:
                current_season = 1
            elif "二" in marker:
                current_season = 2

            plot = clean_text(ws.cell(r, 2).value)
            if not plot or current_season is None or plot.startswith("("):
                continue

            for crop_name, col in crop_col.items():
                area = sol_map.get((year, plot, current_season, crop_name), 0.0)
                if area > 1e-8:
                    ws.cell(r, col).value = round(area, 4)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def evaluate_solution_under_scenarios(
    data: Dict[str, Any],
    params: Dict[str, Any],
    solution_df: pd.DataFrame,
    plan_name: str,
    complement_bonus: float,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    用问题三相关情景评估任意方案，返回情景明细与年度风险摘要。

    加速版说明：
    原版在每个“年份-情景-地块-作物”循环里重复扫描全部 yield/cost 样本，
    这会导致求解器结束后长时间无输出。这里先预计算每年每个情景的平均亩毛利，
    并把方案行转为轻量 list，避免 pandas iterrows 反复开销。
    """
    if solution_df.empty:
        return pd.DataFrame(), pd.DataFrame()

    n_scenarios = int(params["n_scenarios"])
    crop_ids = sorted(data["crop_by_id"].keys())
    bean_ids = data["bean_ids"]

    # 预计算：每年每个情景的平均亩毛利，用于互补收益估计。
    avg_margin_by_year_scenario: Dict[int, np.ndarray] = {}
    for year in YEARS:
        margin_arrays = []
        for cid in crop_ids:
            rel_yields = [arr for (yy, cc, _, _), arr in params["yield_samples"].items() if yy == year and cc == cid]
            rel_costs = [arr for (yy, cc, _, _), arr in params["cost_samples"].items() if yy == year and cc == cid]
            price_arr = params["price_samples"].get((year, cid))
            if not rel_yields or not rel_costs or price_arr is None:
                continue
            mean_yield = np.mean(np.vstack(rel_yields), axis=0)
            mean_cost = np.mean(np.vstack(rel_costs), axis=0)
            margin_arrays.append(np.maximum(0.0, mean_yield * price_arr - mean_cost))
        if margin_arrays:
            avg_margin_by_year_scenario[year] = np.mean(np.vstack(margin_arrays), axis=0)
        else:
            avg_margin_by_year_scenario[year] = np.zeros(n_scenarios, dtype=float)

    # 预计算：plot-year 是否种豆类/非豆类。
    plot_year_bean: Dict[Tuple[str, int], int] = {}
    plot_year_nonbean: Dict[Tuple[str, int], int] = {}
    for row in solution_df.to_dict("records"):
        key = (str(row["地块名"]), int(row["年份"]))
        cid = int(row["作物编号"])
        if cid in bean_ids:
            plot_year_bean[key] = 1
        else:
            plot_year_nonbean[key] = 1

    # 预计算：每年互补面积。这样每个情景只需乘以平均亩毛利。
    complement_area_by_year: Dict[int, float] = {}
    for year in YEARS:
        area_sum = 0.0
        for land in data["lands"]:
            plot = land["plot"]
            prev_bean = data["bean_2023_flag"].get(plot, 0) if year == 2024 else plot_year_bean.get((plot, year - 1), 0)
            cur_nonbean = plot_year_nonbean.get((plot, year), 0)
            if prev_bean and cur_nonbean:
                area_sum += float(land["area"])
        complement_area_by_year[year] = area_sum

    # 将方案按年份转为轻量 records。
    plan_records_by_year: Dict[int, List[Dict[str, Any]]] = {}
    for row in solution_df.to_dict("records"):
        year = int(row["年份"])
        cid = int(row["作物编号"])
        land_type = str(row["地块类型"])
        if "统计季次" in row and clean_text(row.get("统计季次")):
            stat_season = clean_text(row.get("统计季次"))
        elif land_type in ["平旱地", "梯田", "山坡地"]:
            stat_season = "单季"
        elif cid == RICE_ID:
            stat_season = "单季"
        else:
            stat_season = "第一季" if int(row["季次编号"]) == 1 else "第二季"
        plan_records_by_year.setdefault(year, []).append({
            "cid": cid,
            "land_type": land_type,
            "stat_season": stat_season,
            "area": float(row["种植面积/亩"]),
        })

    scenario_rows: List[Dict[str, Any]] = []
    for year in YEARS:
        plan_year = plan_records_by_year.get(year, [])
        if not plan_year:
            continue
        if n_scenarios >= 50:
            print(f"正在评估 {plan_name}：{year} 年，共 {n_scenarios} 个情景...")
        for s in range(n_scenarios):
            production = {cid: 0.0 for cid in crop_ids}
            cost_total = 0.0
            area_total = 0.0
            for row in plan_year:
                cid = row["cid"]
                land_type = row["land_type"]
                stat_season = row["stat_season"]
                area = row["area"]
                yld = params["yield_samples"][(year, cid, land_type, stat_season)][s]
                cost = params["cost_samples"][(year, cid, land_type, stat_season)][s]
                production[cid] += area * yld
                cost_total += area * cost
                area_total += area

            raw_sold = {}
            for cid in crop_ids:
                demand = params["demand_samples"][(year, cid)][s]
                raw_sold[cid] = min(production[cid], demand)

            sold = dict(raw_sold)
            for group, cids in params["groups"].items():
                group_raw = sum(raw_sold[cid] for cid in cids)
                cap = params["group_capacity_samples"][(year, group)][s]
                if group_raw > cap and group_raw > 1e-12:
                    scale = cap / group_raw
                    for cid in cids:
                        sold[cid] *= scale

            revenue = 0.0
            produced_total = 0.0
            sold_total = 0.0
            for cid in crop_ids:
                revenue += sold[cid] * params["price_samples"][(year, cid)][s]
                produced_total += production[cid]
                sold_total += sold[cid]

            complement_value = (
                complement_bonus
                * complement_area_by_year.get(year, 0.0)
                * avg_margin_by_year_scenario[year][s]
            )
            profit = revenue - cost_total + complement_value
            unsold_ratio = 0.0 if produced_total <= 1e-12 else max(0.0, produced_total - sold_total) / produced_total
            scenario_rows.append({
                "方案": plan_name,
                "年份": year,
                "情景编号": s,
                "收入/元": revenue,
                "成本/元": cost_total,
                "互补收益估计/元": complement_value,
                "利润/元": profit,
                "产量/斤": produced_total,
                "销量/斤": sold_total,
                "滞销量/斤": max(0.0, produced_total - sold_total),
                "滞销比例": unsold_ratio,
                "种植面积/亩": area_total,
            })

    scenario_df = pd.DataFrame(scenario_rows)
    if scenario_df.empty:
        return scenario_df, pd.DataFrame()

    summary_rows = []
    for (plan_name_value, year), g in scenario_df.groupby(["方案", "年份"]):
        profits = g["利润/元"].to_numpy(dtype=float)
        p10 = float(np.quantile(profits, 0.10))
        cvar10 = float(profits[profits <= p10].mean()) if np.any(profits <= p10) else p10
        summary_rows.append({
            "方案": plan_name_value,
            "年份": year,
            "平均利润/元": float(np.mean(profits)),
            "利润标准差/元": float(np.std(profits, ddof=1)) if len(profits) > 1 else 0.0,
            "10%分位利润/元": p10,
            "CVaR10/元": cvar10,
            "最差情景利润/元": float(np.min(profits)),
            "平均滞销比例": float(g["滞销比例"].mean()),
        })

    total = scenario_df.groupby(["方案", "情景编号"], as_index=False).agg({"利润/元": "sum", "滞销量/斤": "sum", "产量/斤": "sum"})
    for plan_name_value, g in total.groupby("方案"):
        profits = g["利润/元"].to_numpy(dtype=float)
        p10 = float(np.quantile(profits, 0.10))
        cvar10 = float(profits[profits <= p10].mean()) if np.any(profits <= p10) else p10
        unsold_ratio = float(g["滞销量/斤"].sum() / g["产量/斤"].sum()) if g["产量/斤"].sum() > 1e-12 else 0.0
        summary_rows.append({
            "方案": plan_name_value,
            "年份": "2024-2030合计",
            "平均利润/元": float(np.mean(profits)),
            "利润标准差/元": float(np.std(profits, ddof=1)) if len(profits) > 1 else 0.0,
            "10%分位利润/元": p10,
            "CVaR10/元": cvar10,
            "最差情景利润/元": float(np.min(profits)),
            "平均滞销比例": unsold_ratio,
        })
    return scenario_df, pd.DataFrame(summary_rows)

def build_year_summary(solution_df: pd.DataFrame, sold_df: pd.DataFrame, data: Dict[str, Any], params: Dict[str, Any], status_info: Dict[str, Any]) -> pd.DataFrame:
    rows = []
    for year in YEARS:
        plan = solution_df[solution_df["年份"] == year].copy()
        sold = sold_df[sold_df["年份"] == year].copy()
        cost_total = float((plan["种植面积/亩"] * plan["亩成本/元"]).sum()) if not plan.empty else 0.0
        revenue_total = float((sold["鲁棒销量/斤"] * sold["鲁棒售价/元每斤"]).sum()) if not sold.empty else 0.0
        bean_plots = int(plan[plan["作物编号"].isin(data["bean_ids"])] ["地块名"].nunique()) if not plan.empty else 0
        group_count = int(plan["替代组"].nunique()) if not plan.empty else 0
        rows.append({
            "年份": year,
            "种植记录数": len(plan),
            "种植面积合计/亩": float(plan["种植面积/亩"].sum()) if not plan.empty else 0.0,
            "鲁棒收入/元": revenue_total,
            "鲁棒成本/元": cost_total,
            "鲁棒利润_不含互补收益/元": revenue_total - cost_total,
            "种过豆类的地块数": bean_plots,
            "涉及替代组数量": group_count,
        })
    rows.append({
        "年份": "求解状态",
        "种植记录数": status_info.get("status"),
        "种植面积合计/亩": None,
        "鲁棒收入/元": None,
        "鲁棒成本/元": None,
        "鲁棒利润_不含互补收益/元": status_info.get("robust_objective_with_complement"),
        "种过豆类的地块数": status_info.get("message"),
        "涉及替代组数量": None,
    })
    return pd.DataFrame(rows)


def load_problem2_solution_if_exists(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    df = pd.read_csv(path, encoding="utf-8-sig")
    required = {"年份", "地块名", "地块类型", "季次编号", "作物编号", "作物名称", "作物类型", "种植面积/亩"}
    if not required.issubset(set(df.columns)):
        return pd.DataFrame()
    if "替代组" not in df.columns:
        df["替代组"] = df["作物编号"].astype(int).map(crop_substitution_group)
    return df


def save_outputs(
    base_dir: Path,
    out_dir: Path,
    data: Dict[str, Any],
    params: Dict[str, Any],
    solution_df: pd.DataFrame,
    sold_df: pd.DataFrame,
    status_info: Dict[str, Any],
    scenario_df: pd.DataFrame,
    risk_summary_df: pd.DataFrame,
    comparison_df: pd.DataFrame,
    template_name: str,
    complement_bonus: float,
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    result_path = out_dir / "result3.xlsx"
    template_path = base_dir / "附件3" / template_name
    write_result_template(template_path, result_path, solution_df)

    year_summary_df = build_year_summary(solution_df, sold_df, data, params, status_info)

    group_rows = []
    for group, cids in params["groups"].items():
        group_rows.append({
            "替代组": group,
            "作物编号": ",".join(map(str, cids)),
            "作物名称": "、".join(data["crop_name_by_id"][cid] for cid in cids),
            "建模含义": "组内作物具有市场替代性，销量合计受组市场容量约束",
        })
    group_df = pd.DataFrame(group_rows)

    corr_df = pd.DataFrame(params["group_corr"], index=params["group_names"], columns=params["group_names"]).reset_index().rename(columns={"index": "替代组"})

    model_params_df = pd.DataFrame([
        {"参数": "情景数", "取值": params["n_scenarios"]},
        {"参数": "随机种子", "取值": params["seed"]},
        {"参数": "鲁棒分位数", "取值": params["quantile"]},
        {"参数": "替代组容量系数", "取值": params["group_capacity_alpha"]},
        {"参数": "互补收益系数", "取值": complement_bonus},
        {"参数": "相关性口径", "取值": "需求按替代组相关；价格与需求弱负相关；成本与通胀正相关；产量与天气相关"},
        {"参数": "互补性口径", "取值": "上一年豆类 + 当年非豆类视为轮作互补，加入线性互补收益变量"},
    ])

    with pd.ExcelWriter(out_dir / "problem3_summary.xlsx", engine="openpyxl") as writer:
        year_summary_df.to_excel(writer, sheet_name="年度摘要", index=False)
        risk_summary_df.to_excel(writer, sheet_name="多情景风险评估", index=False)
        comparison_df.to_excel(writer, sheet_name="与问题二比较", index=False)
        solution_df.to_excel(writer, sheet_name="问题三方案长表", index=False)
        sold_df.to_excel(writer, sheet_name="鲁棒销量", index=False)
        group_df.to_excel(writer, sheet_name="替代组设定", index=False)
        corr_df.to_excel(writer, sheet_name="替代组相关矩阵", index=False)
        pd.DataFrame([status_info]).to_excel(writer, sheet_name="求解状态", index=False)
        model_params_df.to_excel(writer, sheet_name="模型参数", index=False)

    solution_df.to_csv(out_dir / "problem3_solution_long.csv", index=False, encoding="utf-8-sig")
    sold_df.to_csv(out_dir / "problem3_sold.csv", index=False, encoding="utf-8-sig")
    scenario_df.to_csv(out_dir / "problem3_evaluation_by_scenario.csv", index=False, encoding="utf-8-sig")
    risk_summary_df.to_csv(out_dir / "problem3_risk_summary.csv", index=False, encoding="utf-8-sig")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="C题问题三：相关性、替代性、互补性鲁棒优化脚本")
    parser.add_argument("--base", type=str, default=".", help="C题目录，默认当前目录")
    parser.add_argument("--att1", type=str, default="附件1.xlsx", help="附件1文件名")
    parser.add_argument("--att2", type=str, default="附件2.xlsx", help="附件2文件名")
    parser.add_argument("--template", type=str, default="result2.xlsx", help="附件3中的同结构模板文件名")
    parser.add_argument("--out", type=str, default="output", help="输出目录")
    parser.add_argument("--scenarios", type=int, default=200, help="随机情景数量；正式跑建议 200-500")
    parser.add_argument("--quantile", type=float, default=0.20, help="鲁棒低分位数，建议 0.15-0.30")
    parser.add_argument("--seed", type=int, default=20240909, help="随机种子")
    parser.add_argument("--time-limit", type=int, default=600, help="MILP 最长求解秒数")
    parser.add_argument("--gap", type=float, default=0.02, help="相对 MIP gap")
    parser.add_argument("--group-capacity-alpha", type=float, default=0.92, help="替代组容量系数，越小表示替代竞争越强")
    parser.add_argument("--complement-bonus", type=float, default=0.03, help="豆类轮作互补收益系数")
    parser.add_argument("--problem2-solution", type=str, default="output/problem2_solution_long.csv", help="问题二长表路径，用于比较；不存在则跳过")
    parser.add_argument("--quiet", action="store_true", help="关闭求解器日志")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    base_dir = Path(args.base).resolve()
    out_dir = base_dir / args.out

    data = load_data(base_dir, args.att1, args.att2)
    params = generate_correlated_parameters(
        data=data,
        n_scenarios=args.scenarios,
        quantile=args.quantile,
        seed=args.seed,
        group_capacity_alpha=args.group_capacity_alpha,
    )

    solution_df, sold_df, status_info = solve_problem3_milp(
        data=data,
        params=params,
        complement_bonus=args.complement_bonus,
        time_limit=args.time_limit,
        gap=args.gap,
        disp=not args.quiet,
    )

    # 关键修正：求解器结束后先落盘核心结果，避免后续风险评估耗时导致长时间没有输出文件。
    out_dir.mkdir(parents=True, exist_ok=True)
    early_result_path = out_dir / "result3.xlsx"
    early_template_path = base_dir / "附件3" / args.template
    write_result_template(early_template_path, early_result_path, solution_df)
    solution_df.to_csv(out_dir / "problem3_solution_long.csv", index=False, encoding="utf-8-sig")
    sold_df.to_csv(out_dir / "problem3_sold.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame([status_info]).to_csv(out_dir / "problem3_solver_status.csv", index=False, encoding="utf-8-sig")
    print(f"求解完成，已先写出核心结果：{early_result_path}")
    print("继续进行多情景风险评估和与问题二比较...")

    scenario_df, risk_summary_df = evaluate_solution_under_scenarios(
        data=data,
        params=params,
        solution_df=solution_df,
        plan_name="问题三方案",
        complement_bonus=args.complement_bonus,
    )

    comparison_parts = [risk_summary_df]
    problem2_path = base_dir / args.problem2_solution
    problem2_df = load_problem2_solution_if_exists(problem2_path)
    if not problem2_df.empty:
        p2_scenario_df, p2_risk_summary_df = evaluate_solution_under_scenarios(
            data=data,
            params=params,
            solution_df=problem2_df,
            plan_name="问题二方案",
            complement_bonus=args.complement_bonus,
        )
        if not p2_scenario_df.empty:
            scenario_df = pd.concat([scenario_df, p2_scenario_df], ignore_index=True)
            comparison_parts.append(p2_risk_summary_df)

    comparison_df = pd.concat(comparison_parts, ignore_index=True) if comparison_parts else pd.DataFrame()

    save_outputs(
        base_dir=base_dir,
        out_dir=out_dir,
        data=data,
        params=params,
        solution_df=solution_df,
        sold_df=sold_df,
        status_info=status_info,
        scenario_df=scenario_df,
        risk_summary_df=risk_summary_df,
        comparison_df=comparison_df,
        template_name=args.template,
        complement_bonus=args.complement_bonus,
    )

    print("\n========== 完成 ==========")
    print(f"求解状态：{status_info['message']}")
    print(f"鲁棒目标值，含互补收益：{status_info['robust_objective_with_complement']}")
    print(f"输出文件：{out_dir / 'result3.xlsx'}")
    print(f"摘要文件：{out_dir / 'problem3_summary.xlsx'}")
    print(f"长表文件：{out_dir / 'problem3_solution_long.csv'}")
    if not problem2_df.empty:
        print(f"已读取问题二方案并完成比较：{problem2_path}")
    else:
        print(f"未找到可比较的问题二长表，已跳过问题二比较：{problem2_path}")


if __name__ == "__main__":
    main()
